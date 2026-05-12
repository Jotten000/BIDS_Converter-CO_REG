"""
Batchverktyg för coregistrering av multimodala PET/MR-data.

Programmet går igenom ett BIDS-liknande dataset med mappar på formen sub-*/ses-*.
För varje session letar programmet efter en T1-bild som referens och coregistrerar
PET till T1-space. Det finns även valfria funktioner för att coregistrera T2,
diffusion/B0 och perfusion till samma T1-referens.

Coregistreringen görs med FSL FLIRT via Nipype. Programmet sparar:
- coregistrerade NIfTI-filer
- transformationsmatriser (.mat)
- JSON-sidecars med metadata
- kommandologgar
- similarity metrics i Excel
- checkerboard-QC-bilder för visuell kvalitetskontroll

GUI:t är byggt med Tkinter och tunga beräkningar körs i separata trådar så att
gränssnittet inte fryser under batch-körningen.
"""


from __future__ import annotations

# Standardbibliotek för filhantering, subprocesser, trådar, tidmätning och JSON.

import json
import re
import shutil
import subprocess
import threading
import queue
import traceback
import os
import shlex
import signal
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

# Externa paket importeras valfritt.
# Om något paket saknas sätts variabeln till None, och programmet kan senare ge ett tydligt felmeddelande när funktionen faktiskt behövs.

try:
    import nibabel as nib
except ImportError:
    nib = None
try:
    import numpy as np
except ImportError:
    np = None

try:
    from nipype.interfaces import fsl
except ImportError:
    fsl = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.comments import Comment
except ImportError:
    Workbook = None
    Font = None
    PatternFill = None
    Alignment = None
    Border = None
    Side = None
    get_column_letter = None
    Table = None
    TableStyleInfo = None
    Comment = None

from packages.Manfred_and_Sofia.File_handling import Conversion_Handler


# Globala inställningar för programmet.
# Här definieras standardvärden för FSL FLIRT, filnamn, QC-inställningar, BIDS-validering och similarity metrics.

APP_TITLE = "Batch coregistrering PET/T2 -> T1 med Nipype + FSL"

DEFAULT_COST = "normmi"
DEFAULT_DOF = 6
OVERWRITE_EXISTING = False

COST_FUNCTION_OPTIONS = (
    "normmi",
    "mutualinfo",
    "corratio",
    "normcorr",
    "leastsq",
)

DEFAULT_DIFFUSION_COST = "normmi"
DEFAULT_PERFUSION_COST = "normmi"

T1_TOKEN = "_T1w"

DIFFUSION_LABEL = "diffusion file"
PERFUSION_LABEL = "perfusion file"

B0_THRESHOLD = 75.0

FSL_COMMAND_TIMEOUT_SECONDS = 2 * 60 * 60  # 2 timmar per FSL-kommando

SIMILARITY_BINS = 64
SIMILARITY_SAMPLE_LIMIT = 500_000
SIMILARITY_METRICS_FILENAME = "coregistration_similarity_metrics.xlsx"

CHECKERBOARD_QC_DIRNAME = "checkerboard_qc"
CHECKERBOARD_BLOCK_SIZE = 32   
CHECKERBOARD_FILENAME_SUFFIX = "_desc-checkerboardQC"
CHECKERBOARD_MARKER = "_desc-checkerboardqc"
REFERENCE_NEGATED = True

VALIDATE_BIDS_BEFORE_RUN_DEFAULT = True
BIDS_VALIDATOR_PRUNE_DERIVATIVES = True
BIDS_VALIDATOR_REPORT_FILENAME = "bids_validator_report.txt"


# Kopplar varje FSL cost-function till en similarity metric som kan användas för kvalitetskontroll efter coregistreringen.
# Exempel: normmi jämförs med Normalized Mutual Information.
COST_FUNCTION_METRIC_MAP = {
    "normmi": {
        "metric_key": "normmi_nmi",
        "metric_name": "Normalized Mutual Information",
        "short_name": "NMI",
        "better_direction": "higher",
        "interpretation": "Högre värde betyder bättre matchning.",
    },
    "mutualinfo": {
        "metric_key": "mutualinfo_mi",
        "metric_name": "Mutual Information",
        "short_name": "MI",
        "better_direction": "higher",
        "interpretation": "Högre värde betyder bättre statistiskt beroende mellan bilderna.",
    },
    "normcorr": {
        "metric_key": "normcorr_ncc",
        "metric_name": "Normalized Cross-Correlation",
        "short_name": "NCC",
        "better_direction": "higher",
        "interpretation": "Högre värde betyder bättre korrelation mellan bilderna.",
    },
    "leastsq": {
        "metric_key": "leastsq_mse",
        "metric_name": "Mean Squared Error",
        "short_name": "MSE",
        "better_direction": "lower",
        "interpretation": "Lägre värde betyder mindre intensitetsskillnad mellan bilderna.",
    },
    "corratio": {
        "metric_key": "correlation_ratio_symmetric",
        "metric_name": "Correlation Ratio",
        "short_name": "Correlation ratio",
        "better_direction": "higher",
        "interpretation": "Högre värde betyder starkare funktionellt samband mellan intensiteterna.",
    },
}

# Kolumner som ska exporteras till Excel-filen med similarity metrics.
SIMILARITY_EXPORT_COLUMNS = [
    "subject",
    "session",
    "coregistration",
    "status",
    "cost_function",
    "similarity_metric_key",
    "similarity_metric_name",
    "similarity_metric_short_name",
    "similarity_metric_value",
    "better_direction",
    "interpretation",
    "dof",
    "metrics_computed",
    "metrics_error",
    "original_voxel_count",
    "used_voxel_count",
    "sampled",
    "mask",
    "histogram_bins",
    "session_dir",
    "moving_file",
    "reference_file",
    "output_file",
    "transform_file",
    "output_json",
]

SIMILARITY_COLUMN_EXPLANATIONS = {
    "subject": "Subject-ID, till exempel sub-001.",
    "session": "Session-ID, till exempel ses-001.",
    "coregistration": "Vilken typ av coregistrering raden gäller, till exempel pet, t2, diffusion eller perfusion.",
    "status": "Status för registreringen, till exempel processed, skipped_existing eller existing_output.",
    "cost_function": "Den FLIRT cost-function som användes vid coregistreringen.",
    "similarity_metric_key": "Intern nyckel för den metric som matchar cost_function.",
    "similarity_metric_name": "Fullständigt namn på den similarity metric som redovisas.",
    "similarity_metric_short_name": "Kort namn på metric, till exempel NMI, MI, NCC eller MSE.",
    "similarity_metric_value": "Själva similarity-värdet för den metric som matchar cost_function.",
    "better_direction": "Anger om högre eller lägre värde är bättre.",
    "interpretation": "Kort tolkning av hur metric-värdet ska läsas.",
    "dof": "Degrees of freedom som användes i FLIRT, till exempel 6.",
    "metrics_computed": "True om similarity metric kunde beräknas, annars False.",
    "metrics_error": "Felmeddelande om metric-beräkningen misslyckades.",
    "original_voxel_count": "Antal voxlar som fanns tillgängliga efter maskning innan eventuell sampling.",
    "used_voxel_count": "Antal voxlar som faktiskt användes i metric-beräkningen.",
    "sampled": "True om bilden var stor och voxel-sampling användes.",
    "mask": "Vilken mask som användes vid metric-beräkningen.",
    "histogram_bins": "Antal bins som användes för histogrambaserade metrics som MI och NMI.",
    "session_dir": "Sökväg till original-sessionens mapp.",
    "moving_file": "Originalbilden som flyttades/coregistrerades.",
    "reference_file": "Referensbilden, vanligtvis T1.",
    "output_file": "Coregistrerad output-bild i T1-space.",
    "transform_file": "FLIRT-transformationsmatrisen.",
    "output_json": "JSON-sidecar för den coregistrerade output-filen.",
}

# Filnamnsmarkörer som används för att känna igen filer som programmet själv har skapat, så att de inte råkar användas som input vid nya körningar.
GENERATED_MARKERS = (
    "_space-t1w_coreg",
    "_desc-coreg",
    "_registered_to_t1",
    "_to_t1",
)

# Intern konfigurationsfil för sparade programvägar.
# Här sparas t.ex. var FSL/flirt finns, så användaren slipper välja igen.
APP_CONFIG_DIR = Path.home() / ".coreg_batch_tool"
APP_CONFIG_FILE = APP_CONFIG_DIR / "settings.json"

FSL_EXECUTABLE_NAMES = ("flirt", "flirt.exe")
FSL_VIEWER_NAMES = ("fsleyes", "fsleyes.exe", "fslview_deprecated", "fslview_deprecated.exe", "fslview", "fslview.exe")


@dataclass
class RuntimeControl:
    """
    Håller reda på aktiva externa processer, till exempel FSL-kommandon.

    Klassen används för att kunna avbryta en pågående batch-körning på ett
    kontrollerat sätt. Alla subprocesser registreras när de startar och tas bort
    när de är klara. Om användaren avbryter körningen kan terminate_all()
    stoppa alla aktiva processer.
    """

    cancel_event: threading.Event
    lock: threading.Lock = field(default_factory=threading.Lock)
    processes: List[subprocess.Popen] = field(default_factory=list)

    def register(self, proc: subprocess.Popen) -> None:
        """
        Registrerar en aktiv subprocess så att den kan stoppas vid avbrott.
        """

        with self.lock:
            self.processes.append(proc)

    def unregister(self, proc: subprocess.Popen) -> None:
        """
        Tar bort en subprocess från listan när den är färdig.
        """
        with self.lock:
            if proc in self.processes:
                self.processes.remove(proc)

    def terminate_all(self) -> None:
        """
        Stoppar alla registrerade processer.

        På Linux/macOS skickas signaler till hela processgruppen så att även
        eventuella barnprocesser stoppas. På Windows används terminate/kill.
        """

        with self.lock:
            procs = list(self.processes)

        for proc in procs:
            if proc.poll() is not None:
                continue
            try:
                if os.name != "nt":
                    os.killpg(proc.pid, signal.SIGTERM)
                else:
                    proc.terminate()
            except Exception:
                pass

        for proc in procs:
            if proc.poll() is not None:
                continue
            try:
                proc.wait(timeout=3)
            except subprocess.TimeoutExpired:
                try:
                    if os.name != "nt":
                        os.killpg(proc.pid, signal.SIGKILL)
                    else:
                        proc.kill()
                except Exception:
                    pass


def get_available_cpu_count() -> int:
    """
    Returnerar antal tillgängliga logiska CPU-kärnor.
    os.cpu_count() kan returnera None, så vi faller tillbaka till 1.
    """

    return os.cpu_count() or 1


def get_recommended_worker_count() -> int:
    """
    Rekommenderar antal parallella workers.

    För coregistrering är det ofta bättre att lämna minst en CPU-kärna fri,
    eftersom FSL och systemet annars kan bli långsamt.
    """

    cores = get_available_cpu_count()

    if cores <= 2:
        return 1

    return min(cores - 1, 4)


def get_worker_options(max_cores: Optional[int] = None) -> Tuple[str, ...]:
    """
    Skapar val till comboboxen baserat på hur många CPU-kärnor datorn har.
    """
    
    cores = max_cores or get_available_cpu_count()

    base_values = (1, 2, 4, 6, 8, 16, 24, 32, 48, 64, 80, 96, 128)

    values = sorted({
        value for value in base_values
        if value <= cores
    } | {cores})

    return tuple(str(value) for value in values)


def is_nifti_file(path: str) -> bool:
    """
    Kontrollerar om en fil är en NIfTI-fil, antingen .nii eller .nii.gz.
    """

    lower = path.lower()
    return lower.endswith(".nii") or lower.endswith(".nii.gz")


def ensure_python_dependencies() -> None:
    """
    Kontrollerar att alla Python-paket som behövs är installerade.

    Programmet använder denna funktion innan körning, så att användaren får
    ett tydligt felmeddelande om exempelvis nibabel, numpy, nipype eller
    openpyxl saknas.
    """

    missing = []
    if nib is None:
        missing.append("nibabel")
    if np is None:
        missing.append("numpy")
    if fsl is None:
        missing.append("nipype")
    if Workbook is None:
        missing.append("openpyxl")

    if missing:
        raise RuntimeError(
            "Följande Python-paket saknas: "
            + ", ".join(missing)
            + ". Installera med: pip install "
            + " ".join(missing)
        )


def find_executable(name: str) -> Optional[str]:
    """
    Letar efter ett körbart program först i PATH, sedan i PyInstaller-bundlen.
    """

    found = shutil.which(name)
    if found:
        return found

    names = (name,)

    if os.name == "nt" and not name.lower().endswith(".exe"):
        names = (name, name + ".exe")

    return find_bundled_executable(names)

def find_bundled_executable(names: Tuple[str, ...]) -> Optional[str]:
    """
    Letar efter ett externt program som PyInstaller har packat med,
    t.ex. deno eller dcm2niix.
    """

    candidate_dirs: List[Path] = []

    if getattr(sys, "frozen", False):
        exe_dir = Path(sys.executable).resolve().parent

        candidate_dirs.extend([
            exe_dir,
            exe_dir.parent / "Frameworks",
            exe_dir.parent / "Resources",
        ])

        meipass = getattr(sys, "_MEIPASS", None)
        if meipass:
            meipass_path = Path(meipass).resolve()
            candidate_dirs.extend([
                meipass_path,
                meipass_path / "bin",
                meipass_path.parent / "Frameworks",
                meipass_path.parent / "Resources",
            ])

    candidate_dirs.append(Path.cwd())

    seen = set()
    unique_dirs = []

    for folder in candidate_dirs:
        try:
            folder = folder.resolve()
        except Exception:
            continue

        if folder in seen:
            continue

        seen.add(folder)

        if folder.exists() and folder.is_dir():
            unique_dirs.append(folder)

    for folder in unique_dirs:
        for name in names:
            candidate = folder / name
            if is_executable_file(candidate):
                return str(candidate)

    # Extra fallback: sök rekursivt i appens närmaste mappar.
    for folder in unique_dirs:
        for name in names:
            try:
                for candidate in folder.rglob(name):
                    if is_executable_file(candidate):
                        return str(candidate)
            except Exception:
                pass

    return None


def load_app_config() -> Dict[str, Any]:
    """
    Läser programmets interna konfigurationsfil.

    Används för att komma ihåg t.ex. var FSL/flirt finns.
    """

    if not APP_CONFIG_FILE.exists():
        return {}

    try:
        with open(APP_CONFIG_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)

        if isinstance(data, dict):
            return data

    except Exception:
        pass

    return {}


def save_app_config(config: Dict[str, Any]) -> None:
    """
    Sparar programmets interna konfiguration.
    """

    APP_CONFIG_DIR.mkdir(parents=True, exist_ok=True)

    with open(APP_CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(config, f, indent=2, ensure_ascii=False)


def is_executable_file(path: Path) -> bool:
    """
    Kontrollerar om path är en körbar fil.
    På Windows räcker det oftast att filen finns.
    """

    if not path.is_file():
        return False

    if os.name == "nt":
        return True

    return os.access(path, os.X_OK)


def prepend_to_path(directory: Path) -> None:
    """
    Lägger en mapp först i PATH om den inte redan finns där.
    """

    directory_str = str(directory)
    current_path = os.environ.get("PATH", "")
    parts = current_path.split(os.pathsep) if current_path else []

    if directory_str not in parts:
        os.environ["PATH"] = directory_str + (
            os.pathsep + current_path if current_path else ""
        )


def find_executable_in_folder(folder: Path, names: Tuple[str, ...]) -> Optional[Path]:
    """
    Letar efter ett program antingen direkt i folder eller i folder/bin.

    Detta gör att användaren kan välja antingen:
    - FSL-mappen, t.ex. /usr/local/fsl
    - FSL bin-mappen, t.ex. /usr/local/fsl/bin
    - själva flirt-filen
    """

    folder = folder.expanduser()

    candidate_dirs = [
        folder,
        folder / "bin",
    ]

    for candidate_dir in candidate_dirs:
        for name in names:
            candidate = candidate_dir / name
            if is_executable_file(candidate):
                return candidate.resolve()

    return None


def running_on_windows() -> bool:
    """
    True om programmet körs nativt på Windows.

    OBS: WSL räknas inte som Windows här, eftersom os.name då normalt är 'posix'.
    """
    return os.name == "nt" or sys.platform.startswith("win")


WINDOWS_FSL_DISABLED_MESSAGE = (
    "Programmet körs nativt i Windows.\n\n"
    "FSL/FSLeyes är därför avstängt i denna version.\n"
    "Programmet kommer inte fråga efter FSL-sökväg, och det går inte att öppna "
    "NIfTI-bilder i FSL från Windows-läget.\n\n"
    "Kör programmet i WSL, Linux eller macOS om du vill använda FSL-coregistrering "
    "eller öppna bilder i FSLeyes."
)


def raise_if_windows_fsl_feature(feature_name: str) -> None:
    """
    Stoppar funktioner som kräver FSL/FSLeyes när programmet körs på Windows.
    """
    if running_on_windows():
        raise RuntimeError(
            f"{feature_name} kräver FSL/FSLeyes och är avstängt när programmet körs nativt i Windows.\n\n"
            "Kör programmet i WSL, Linux eller macOS för den funktionen."
        )


def configure_fsl_from_flirt(flirt_path: Path) -> None:
    """
    Ställer in miljön så att FSL-kommandon fungerar.

    Nipype bygger kommandon som börjar med t.ex. 'flirt'.
    Därför måste FSL:s bin-mapp finnas i PATH.
    """

    flirt_path = flirt_path.resolve()
    fsl_bin = flirt_path.parent

    prepend_to_path(fsl_bin)

    if fsl_bin.name.lower() == "bin":
        os.environ["FSLDIR"] = str(fsl_bin.parent)

    os.environ["FSLOUTPUTTYPE"] = "NIFTI_GZ"

    # Informera Nipype om output-typen om nipype är importerat.
    try:
        if fsl is not None:
            fsl.FSLCommand.set_default_output_type("NIFTI_GZ")
    except Exception:
        pass


def remember_fsl_flirt_path(flirt_path: Path) -> None:
    """
    Sparar FSL/flirt-sökvägen i programmets interna config.
    """
    
    flirt_path = flirt_path.resolve()

    config = load_app_config()
    config["fsl_flirt_executable"] = str(flirt_path)

    if flirt_path.parent.name.lower() == "bin":
        config["fsl_bin_dir"] = str(flirt_path.parent)
        config["fsl_dir"] = str(flirt_path.parent.parent)

    save_app_config(config)


def discover_fsl_flirt() -> Optional[Path]:
    """
    Försöker hitta FSL flirt automatiskt.

    Prioritet:
    1. Sparad intern config
    2. PATH
    3. FSLDIR-miljövariabel
    4. Vanliga installationsmappar
    """

    # 1. Sparad config
    config = load_app_config()

    saved_flirt = config.get("fsl_flirt_executable")
    if saved_flirt:
        path = Path(saved_flirt).expanduser()
        if is_executable_file(path):
            path = path.resolve()
            configure_fsl_from_flirt(path)
            return path

    saved_fsl_dir = config.get("fsl_dir")
    if saved_fsl_dir:
        path = find_executable_in_folder(Path(saved_fsl_dir), FSL_EXECUTABLE_NAMES)
        if path is not None:
            configure_fsl_from_flirt(path)
            remember_fsl_flirt_path(path)
            return path

    saved_fsl_bin = config.get("fsl_bin_dir")
    if saved_fsl_bin:
        path = find_executable_in_folder(Path(saved_fsl_bin), FSL_EXECUTABLE_NAMES)
        if path is not None:
            configure_fsl_from_flirt(path)
            remember_fsl_flirt_path(path)
            return path

    # 2. PATH
    for name in FSL_EXECUTABLE_NAMES:
        found = shutil.which(name)
        if found:
            path = Path(found).resolve()
            configure_fsl_from_flirt(path)
            remember_fsl_flirt_path(path)
            return path

    # 3. FSLDIR
    fsldir = os.environ.get("FSLDIR")
    if fsldir:
        path = find_executable_in_folder(Path(fsldir), FSL_EXECUTABLE_NAMES)
        if path is not None:
            configure_fsl_from_flirt(path)
            remember_fsl_flirt_path(path)
            return path

    # 4. Vanliga platser på Linux/macOS/WSL
    common_fsl_dirs = [
        Path("/usr/local/fsl"),
        Path("/usr/share/fsl"),
        Path("/opt/fsl"),
        Path("/opt/fsl-6.0"),
        Path.home() / "fsl",
    ]

    for folder in common_fsl_dirs:
        path = find_executable_in_folder(folder, FSL_EXECUTABLE_NAMES)
        if path is not None:
            configure_fsl_from_flirt(path)
            remember_fsl_flirt_path(path)
            return path

    return None


def prompt_user_for_fsl_location(parent: tk.Misc) -> Optional[Path]:
    """
    Frågar användaren var FSL/flirt finns och returnerar flirt-sökvägen.

    Användaren kan välja antingen FSL-mappen eller själva flirt-filen.
    """

    while True:
        answer = messagebox.askyesnocancel(
            "Välj FSL",
            "Välj var FSL-kommandot 'flirt' finns.\n\n"
            "Ja = välj FSL-mappen, till exempel /usr/local/fsl\n"
            "Nej = välj själva flirt-filen, till exempel /usr/local/fsl/bin/flirt\n"
            "Avbryt = stoppa körningen",
            parent=parent,
        )

        if answer is None:
            return None

        if answer is True:
            selected_dir = filedialog.askdirectory(
                title="Välj FSL-mappen eller FSL bin-mappen",
                initialdir=str(Path.home()),
                parent=parent,
            )

            if not selected_dir:
                return None

            flirt_path = find_executable_in_folder(
                Path(selected_dir),
                FSL_EXECUTABLE_NAMES,
            )

            if flirt_path is not None:
                return flirt_path

            retry = messagebox.askretrycancel(
                "Hittade inte flirt",
                "Kunde inte hitta 'flirt' i den valda mappen.\n\n"
                "Välj antingen FSL-mappen, t.ex. /usr/local/fsl,\n"
                "eller FSL:s bin-mapp, t.ex. /usr/local/fsl/bin.",
                parent=parent,
            )

            if not retry:
                return None

        else:
            selected_file = filedialog.askopenfilename(
                title="Välj FSL-kommandot flirt",
                initialdir=str(Path.home()),
                parent=parent,
            )

            if not selected_file:
                return None

            flirt_path = Path(selected_file).expanduser()

            if flirt_path.name.lower() not in FSL_EXECUTABLE_NAMES:
                retry = messagebox.askretrycancel(
                    "Fel fil vald",
                    "Filen du valde heter inte 'flirt'.\n\n"
                    "Välj filen som vanligtvis ligger här:\n"
                    "/usr/local/fsl/bin/flirt",
                    parent=parent,
                )

                if not retry:
                    return None

                continue

            if not is_executable_file(flirt_path):
                retry = messagebox.askretrycancel(
                    "Filen är inte körbar",
                    "Den valda filen verkar inte vara körbar.\n\n"
                    "Kontrollera att du valt rätt FSL/flirt-fil.",
                    parent=parent,
                )

                if not retry:
                    return None

                continue

            return flirt_path.resolve()


def ensure_fsl_available_gui(parent: tk.Misc) -> str:
    """
    Säkerställer att FSL/flirt finns.

    Om FSL hittas automatiskt används det.

    Annars får användaren välja FSL, och valet sparas för framtida körningar.

    Om programmet körs nativt på Windows frågar vi inte efter FSL.
    """
    if running_on_windows():
        raise RuntimeError(WINDOWS_FSL_DISABLED_MESSAGE)

    flirt_path = discover_fsl_flirt()

    if flirt_path is None:
        selected = prompt_user_for_fsl_location(parent)

        if selected is None:
            raise RuntimeError(
                "FSL/flirt hittades inte och ingen sökväg valdes. "
                "Coregistreringen kan inte starta utan FSL."
            )

        configure_fsl_from_flirt(selected)
        remember_fsl_flirt_path(selected)
        flirt_path = selected

    if shutil.which("flirt") is None and shutil.which("flirt.exe") is None:
        raise RuntimeError(
            "FSL/flirt valdes, men kommandot kunde ändå inte hittas i PATH.\n\n"
            f"Vald flirt-sökväg:\n{flirt_path}\n\n"
            "Kontrollera att du har valt rätt FSL-mapp eller FSL/bin-mapp."
        )

    return str(flirt_path)


def ensure_fsl_available() -> None:
    """
    Icke-GUI-version. Används om FSL behöver kontrolleras utanför Tkinter.
    """

    if running_on_windows():
        raise RuntimeError(WINDOWS_FSL_DISABLED_MESSAGE)

    flirt_path = discover_fsl_flirt()

    if flirt_path is None:
        raise RuntimeError(
            "FSL verkar inte vara installerat eller så finns 'flirt' inte i PATH. "
            "Starta programmet via GUI så kan du välja FSL manuellt."
        )


def find_fsl_viewer_executable() -> Optional[str]:
    """
    Letar efter en installerad FSL-bildvisare.

    Söker först i PATH. Om FSL har sparats internt söker den även i FSL:s bin-mapp.
    På Windows returneras None direkt.
    """

    if running_on_windows():
        return None

    for name in FSL_VIEWER_NAMES:
        exe = find_executable(name)
        if exe is not None:
            return exe

    config = load_app_config()

    for key in ("fsl_bin_dir", "fsl_dir"):
        value = config.get(key)
        if not value:
            continue

        folder = Path(value).expanduser()
        viewer = find_executable_in_folder(folder, FSL_VIEWER_NAMES)

        if viewer is not None:
            return str(viewer)

    flirt_path = discover_fsl_flirt()
    if flirt_path is not None:
        viewer = find_executable_in_folder(flirt_path.parent, FSL_VIEWER_NAMES)
        if viewer is not None:
            return str(viewer)

    return None


def read_text_file(path: str) -> str:
    """
    Läser in en textfil som UTF-8 och returnerar innehållet som en sträng.
    """

    with open(path, "r", encoding="utf-8") as f:
        return f.read()


# Funktioner för att läsa och skriva JSON.
# load_json_lenient försöker även reparera enklare fel i JSON-mallar,
# till exempel saknade värden som ersätts med null.
def repair_json_template_text(text: str) -> str:
    """
    Försöker reparera enkla fel i JSON-liknande text.

    Funktionen ersätter tomma JSON-värden med null och tar bort avslutande
    kommatecken före hakparentes.
    """

    repaired = text.strip()
    repaired = re.sub(r":\s*(,|\})", r": null\1", repaired)
    repaired = re.sub(r",\s*\]", "]", repaired)
    return repaired

def load_json_lenient(path: str) -> Dict[str, Any]:
    """
    Läser en JSON-fil och accepterar vissa enklare mallfel.

    Om vanlig JSON-läsning misslyckas försöker funktionen först reparera texten.
    Returnerar alltid ett dictionary-objekt.
    """

    raw = read_text_file(path)
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        repaired = repair_json_template_text(raw)
        data = json.loads(repaired)

    if not isinstance(data, dict):
        raise ValueError(f"JSON-filen måste innehålla ett objekt: {path}")
    return data

def save_json(path: str, data: Dict[str, Any]) -> None:
    """
    Sparar ett dictionary-objekt som formaterad JSON med UTF-8-kodning.
    """

    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def format_duration(seconds: float) -> str:
    """
    Formaterar en tidslängd i sekunder till en läsbar text.

    Exempel: 3661 sekunder blir "1 h 1 min 1 s".
    """

    seconds = max(0, int(round(seconds)))

    hours, remainder = divmod(seconds, 3600)
    minutes, secs = divmod(remainder, 60)

    if hours > 0:
        return f"{hours} h {minutes} min {secs} s"
    if minutes > 0:
        return f"{minutes} min {secs} s"
    return f"{secs} s"


def safe_filename_token(value: str) -> str:
    """
    Gör text säker att använda i filnamn, t.ex. cost-normmi.
    """

    cleaned = re.sub(r"[^A-Za-z0-9_.-]+", "-", value.strip())
    cleaned = cleaned.strip("._-")
    return cleaned or "unknown"


def normalize_value_for_json(value: Any) -> Any:
    """
    Konverterar NumPy-typer till vanliga Python-typer.

    Detta behövs eftersom NumPy-arrayer och NumPy-tal inte alltid kan sparas
    direkt i JSON.
    """

    if np is not None:
        if isinstance(value, np.floating):
            return float(value)
        if isinstance(value, np.integer):
            return int(value)
        if isinstance(value, np.ndarray):
            return value.tolist()
    return value


def affine_to_dicom_iop(affine: np.ndarray) -> Optional[List[float]]:
    """
    Beräknar DICOM-fältet ImageOrientationPatient från NIfTI-affinen.

    NIfTI använder oftast RAS-koordinater medan DICOM använder LPS.
    Därför konverteras riktningarna från RAS till LPS innan de sparas.
    """

    mat = affine[:3, :3]
    if mat.shape != (3, 3):
        return None

    col_x = mat[:, 0]
    col_y = mat[:, 1]
    nx = np.linalg.norm(col_x)
    ny = np.linalg.norm(col_y)
    if nx == 0 or ny == 0:
        return None

    row_dir_ras = col_x / nx
    col_dir_ras = col_y / ny
    ras_to_lps = np.array([-1.0, -1.0, 1.0])

    row_dir_lps = row_dir_ras * ras_to_lps
    col_dir_lps = col_dir_ras * ras_to_lps
    return [float(v) for v in np.concatenate([row_dir_lps, col_dir_lps])]


def nifti_info(path: str) -> Dict[str, Any]:
    """
    Läser grundläggande metadata från en NIfTI-fil.

    Informationen används senare i JSON-sidecar-filerna för att dokumentera
    bildens form, voxelstorlek, datatyp, affine-matris och orientering.
    """

    img = nib.load(path)
    hdr = img.header
    zooms = hdr.get_zooms()

    info: Dict[str, Any] = {
        "source_file": str(path),
        "shape": [int(x) for x in img.shape],
        "voxel_sizes": [float(x) for x in zooms[: len(img.shape)]],
        "dtype": str(img.get_data_dtype()),
        "affine": img.affine.tolist(),
        "axis_codes": list(nib.aff2axcodes(img.affine)),
    }

    if np is not None and len(img.shape) > 0:
        data = np.asanyarray(img.dataobj)
        finite = data[np.isfinite(data)] if data.size > 0 else np.array([])
        if finite.size > 0:
            info["data_min"] = float(np.min(finite))
            info["data_max"] = float(np.max(finite))

    iop = affine_to_dicom_iop(img.affine)
    if iop:
        info["ImageOrientationPatientDICOM"] = iop

    return info


def _to_3d_metric_data(path: str) -> Tuple[np.ndarray, str]:
    """
    Returnerar 3D-data för similarity-beräkning utan att ladda hela 4D-bilder i RAM.

    3D-bild: läses som float32.
    4D-bild: endast första volymen används.
    """

    img = nib.load(path)
    shape = img.shape

    if len(shape) == 3:
        data = np.asarray(img.dataobj, dtype=np.float32)
        return data, "3D"

    if len(shape) == 4:
        if shape[3] < 1:
            raise RuntimeError(f"4D-bilden har inga volymer: {path}")

        data = np.asarray(img.dataobj[..., 0], dtype=np.float32)
        return data, "4D_first_volume"

    raise RuntimeError(f"Kan inte beräkna similarity metric för {len(shape)}D-bild: {path}")


def _entropy_from_probabilities(p: np.ndarray) -> Optional[float]:
    """
    Beräknar entropi från en sannolikhetsfördelning.

    Nollvärden tas bort innan logaritmen beräknas för att undvika numeriska fel.
    """

    p = p[p > 0]
    if p.size == 0:
        return None
    return float(-np.sum(p * np.log(p)))


def _mutual_information_metrics(
    fixed: np.ndarray,
    moving: np.ndarray,
    bins: int = SIMILARITY_BINS,
) -> Dict[str, Optional[float]]:
    """
    Beräknar Mutual Information och Normalized Mutual Information.

    Dessa mått används ofta vid multimodal registrering, eftersom PET, T1 och T2
    kan ha olika intensitetsskalor men ändå innehålla anatomiskt relaterad
    information.
    """

    hist_2d, _, _ = np.histogram2d(fixed, moving, bins=bins)

    total = np.sum(hist_2d)
    if total <= 0:
        return {
            "mi": None,
            "nmi": None,
            "normalized_mi": None,
            "fixed_entropy": None,
            "moving_entropy": None,
            "joint_entropy": None,
        }

    pxy = hist_2d / total
    px = np.sum(pxy, axis=1)
    py = np.sum(pxy, axis=0)

    hx = _entropy_from_probabilities(px)
    hy = _entropy_from_probabilities(py)
    hxy = _entropy_from_probabilities(pxy.ravel())

    px_py = px[:, None] * py[None, :]
    nz = pxy > 0

    mi = float(np.sum(pxy[nz] * np.log(pxy[nz] / px_py[nz])))

    if hxy is not None and hxy > 0 and hx is not None and hy is not None:
        nmi = float((hx + hy) / hxy)
    else:
        nmi = None

    if hx is not None and hy is not None and (hx + hy) > 0:
        normalized_mi = float(2.0 * mi / (hx + hy))
    else:
        normalized_mi = None

    return {
        "mi": mi,
        "nmi": nmi,
        "normalized_mi": normalized_mi,
        "fixed_entropy": hx,
        "moving_entropy": hy,
        "joint_entropy": hxy,
    }


def _correlation_ratio(
    categories_source: np.ndarray,
    measured_values: np.ndarray,
    bins: int = SIMILARITY_BINS,
) -> Optional[float]:
    """
    Beräknar en approximativ correlation ratio.

    Måttet beskriver hur väl intensiteter i en bild kan förklara intensiteter
    i den andra bilden. Det är användbart när relationen inte nödvändigtvis är
    linjär.
    
    Approximation av correlation ratio eta^2.
    Returnerar värde mellan ungefär 0 och 1.
    """

    if categories_source.size < 2:
        return None

    if np.all(categories_source == categories_source[0]):
        return None

    if np.all(measured_values == measured_values[0]):
        return None

    edges = np.histogram_bin_edges(categories_source, bins=bins)
    bin_ids = np.digitize(categories_source, edges[1:-1], right=False)

    global_mean = float(np.mean(measured_values))
    denominator = float(np.sum((measured_values - global_mean) ** 2))

    if denominator <= 0:
        return None

    numerator = 0.0

    for bin_id in range(bins):
        values_in_bin = measured_values[bin_ids == bin_id]
        if values_in_bin.size == 0:
            continue

        local_mean = float(np.mean(values_in_bin))
        numerator += values_in_bin.size * ((local_mean - global_mean) ** 2)

    return float(numerator / denominator)


def compute_similarity_metrics(
    reference_file: str,
    registered_file: str,
    bins: int = SIMILARITY_BINS,
    sample_limit: int = SIMILARITY_SAMPLE_LIMIT,
) -> Dict[str, Any]:
    """
    Beräknar kvalitetsmått mellan T1-referensen och den coregistrerade bilden.

    Funktionen används efter registrering för att ge ett numeriskt mått på hur
    väl bilderna matchar. Den beräknar bland annat MSE, RMSE, NCC, MI, NMI och
    correlation ratio.

    För att undvika att bakgrundsvoxlar dominerar resultatet används i första
    hand en mask där båda bilderna är finita och icke-noll. Om för få voxlar
    finns används en enklare finite-mask som fallback.


    Beräknar flera similarity metrics mellan T1-reference och coregistrerad output.

    reference_file = T1
    registered_file = output från coregistrering, alltså redan i T1-space
    """

    if nib is None or np is None:
        raise RuntimeError("Similarity metrics kräver nibabel och numpy.")

    fixed_data, fixed_kind = _to_3d_metric_data(reference_file)
    moving_data, moving_kind = _to_3d_metric_data(registered_file)

    if fixed_data.shape != moving_data.shape:
        raise RuntimeError(
            "Reference och coregistrerad bild har olika shape. "
            f"Reference: {fixed_data.shape}, registered: {moving_data.shape}"
        )

    finite_mask = np.isfinite(fixed_data) & np.isfinite(moving_data)

    # Undvik att bakgrund dominerar metriken.
    mask = finite_mask & (fixed_data != 0) & (moving_data != 0)
    mask_name = "finite_and_both_nonzero"

    # Fallback om det blev för få voxlar.
    if int(np.sum(mask)) < 10:
        mask = finite_mask
        mask_name = "finite_only"

    fixed = fixed_data[mask].ravel()
    moving = moving_data[mask].ravel()

    original_voxel_count = int(fixed.size)

    if original_voxel_count < 2:
        raise RuntimeError(
            f"För få voxlar för similarity metrics: {original_voxel_count}"
        )

    sampled = False

    # Deterministisk sampling för stora bilder.
    if sample_limit > 0 and fixed.size > sample_limit:
        idx = np.linspace(0, fixed.size - 1, sample_limit).astype(np.int64)
        fixed = fixed[idx]
        moving = moving[idx]
        sampled = True

    diff = fixed - moving

    ssd = float(np.sum(diff ** 2))
    mse = float(np.mean(diff ** 2))
    rmse = float(np.sqrt(mse))
    mae = float(np.mean(np.abs(diff)))

    fixed_mean = float(np.mean(fixed))
    moving_mean = float(np.mean(moving))

    fixed_centered = fixed - fixed_mean
    moving_centered = moving - moving_mean

    ncc_denominator = np.sqrt(
        np.sum(fixed_centered ** 2) * np.sum(moving_centered ** 2)
    )

    if ncc_denominator > 0:
        ncc = float(np.sum(fixed_centered * moving_centered) / ncc_denominator)
    else:
        ncc = None

    mi_metrics = _mutual_information_metrics(
        fixed=fixed,
        moving=moving,
        bins=bins,
    )

    corr_ratio_fixed_to_moving = _correlation_ratio(
        categories_source=fixed,
        measured_values=moving,
        bins=bins,
    )

    corr_ratio_moving_to_fixed = _correlation_ratio(
        categories_source=moving,
        measured_values=fixed,
        bins=bins,
    )

    corr_values = [
        v for v in (corr_ratio_fixed_to_moving, corr_ratio_moving_to_fixed)
        if v is not None
    ]

    if corr_values:
        corr_ratio_symmetric = float(np.mean(corr_values))
    else:
        corr_ratio_symmetric = None

    return {
        "computed": True,
        "reference_file": str(reference_file),
        "registered_file": str(registered_file),
        "reference_data_kind": fixed_kind,
        "registered_data_kind": moving_kind,
        "mask": mask_name,
        "original_voxel_count": original_voxel_count,
        "used_voxel_count": int(fixed.size),
        "sampled": bool(sampled),
        "histogram_bins": int(bins),

        # Namn som matchar FSL-cost-funktionerna.
        "leastsq_mse": mse,
        "normcorr_ncc": ncc,
        "mutualinfo_mi": mi_metrics["mi"],
        "normmi_nmi": mi_metrics["nmi"],
        "correlation_ratio_fixed_to_registered": corr_ratio_fixed_to_moving,
        "correlation_ratio_registered_to_fixed": corr_ratio_moving_to_fixed,
        "correlation_ratio_symmetric": corr_ratio_symmetric,

        # Extra QC-mått som sparas i JSON men inte visas som huvudkolumner i Excel.
        "ssd": ssd,
        "mse": mse,
        "rmse": rmse,
        "mae": mae,
        "ncc": ncc,
        "mi": mi_metrics["mi"],
        "nmi": mi_metrics["nmi"],
        "normalized_mi": mi_metrics["normalized_mi"],
        "fixed_entropy": mi_metrics["fixed_entropy"],
        "moving_entropy": mi_metrics["moving_entropy"],
        "joint_entropy": mi_metrics["joint_entropy"],
    }


def compute_similarity_metrics_safe(
    reference_file: str,
    registered_file: str,
) -> Dict[str, Any]:
    """
    Wrapper så att hela batchen inte kraschar om metric-beräkning misslyckas.
    """

    try:
        return compute_similarity_metrics(
            reference_file=reference_file,
            registered_file=registered_file,
        )
    except Exception as exc:
        return {
            "computed": False,
            "reference_file": str(reference_file),
            "registered_file": str(registered_file),
            "error": str(exc),
        }


def similarity_metric_info_for_cost(cost_function: Optional[str]) -> Dict[str, str]:
    """
    Hämtar information om vilken similarity metric som hör till en cost-function.

    Om cost-function är okänd returneras standardinformation som markerar att
    ingen matchande metric kunde väljas.
    """

    cost = (cost_function or "").strip().lower()
    return COST_FUNCTION_METRIC_MAP.get(
        cost,
        {
            "metric_key": "unknown",
            "metric_name": "Unknown metric",
            "short_name": "unknown",
            "better_direction": "unknown",
            "interpretation": "Okänd cost-function. Ingen matchande similarity metric kunde väljas.",
        },
    )


def similarity_row_from_registration(
    subject: str,
    session: str,
    session_dir: str,
    coregistration: str,
    reg: Dict[str, Any],
) -> Dict[str, Any]:
    """
    Skapar en Excel-rad från resultatet av en enskild registrering.

    Funktionen plockar ut subject, session, cost-function, similarity metric,
    outputfiler och övrig metadata i ett standardiserat format.
    """

    metrics = reg.get("similarity_metrics", {})
    cost_function = reg.get("cost_function")

    metric_info = similarity_metric_info_for_cost(cost_function)
    metric_key = metric_info["metric_key"]

    return {
        "subject": subject,
        "session": session,
        "coregistration": coregistration,
        "status": reg.get("status"),
        "cost_function": cost_function,
        "similarity_metric_key": metric_key,
        "similarity_metric_name": metric_info["metric_name"],
        "similarity_metric_short_name": metric_info["short_name"],
        "similarity_metric_value": metrics.get(metric_key),
        "better_direction": metric_info["better_direction"],
        "interpretation": metric_info["interpretation"],
        "dof": reg.get("dof"),
        "metrics_computed": metrics.get("computed"),
        "metrics_error": metrics.get("error"),
        "original_voxel_count": metrics.get("original_voxel_count"),
        "used_voxel_count": metrics.get("used_voxel_count"),
        "sampled": metrics.get("sampled"),
        "mask": metrics.get("mask"),
        "histogram_bins": metrics.get("histogram_bins"),
        "session_dir": session_dir,
        "moving_file": reg.get("moving_file"),
        "reference_file": reg.get("reference_file"),
        "output_file": reg.get("output_file"),
        "transform_file": reg.get("transform_file"),
        "output_json": reg.get("output_json"),
    }


def collect_similarity_metric_rows(
    session_results: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """
    Samlar similarity metric-rader från alla processade sessioner.

    Varje PET-, T2-, diffusion- och perfusionsregistrering blir en egen rad
    i Excel-exporten om resultatet finns.
    """

    rows: List[Dict[str, Any]] = []

    for session_result in session_results:
        subject = session_result.get("subject")
        session = session_result.get("session")
        session_dir = session_result.get("session_dir")

        for coregistration in ("pet", "t2", "diffusion", "perfusion"):
            reg = session_result.get(coregistration)

            if not isinstance(reg, dict):
                continue

            row = similarity_row_from_registration(
                subject=str(subject),
                session=str(session),
                session_dir=str(session_dir),
                coregistration=coregistration,
                reg=reg,
            )
            rows.append(row)

    return rows


def excel_safe_value(value: Any) -> Any:
    """
    Gör ett värde säkert att skriva till Excel.

    Långa texter kortas ned eftersom Excel har en maxgräns för cellinnehåll.
    """

    if value is None:
        return None
    if isinstance(value, (int, float, bool)):
        return value
    text = str(value)
    if len(text) > 32000:
        return text[:32000] + "[truncated]"
    return text


def write_similarity_metrics_xlsx(
    path: str,
    rows: List[Dict[str, Any]],
) -> None:
    """
    Skapar en formaterad Excel-fil med similarity metrics.

    Excel-filen innehåller en huvudflik med metrics per registrering och en
    separat flik med förklaringar till varje kolumn. Rubrikerna formateras,
    filter aktiveras och kolumnbredder anpassas automatiskt.
    """

    if Workbook is None:
        raise RuntimeError(
            "Excel-export kräver openpyxl. Installera med: pip install openpyxl"
        )

    wb = Workbook()

    ws = wb.active
    ws.title = "Similarity metrics"

    columns = SIMILARITY_EXPORT_COLUMNS
    ws.append(columns)

    for row in rows:
        ws.append([excel_safe_value(row.get(col)) for col in columns])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    thin_gray = Side(style="thin", color="D9D9D9")
    border = Border(
        left=thin_gray,
        right=thin_gray,
        top=thin_gray,
        bottom=thin_gray,
    )

    for cell in ws[1]:
        column_name = str(cell.value)

        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

        explanation = SIMILARITY_COLUMN_EXPLANATIONS.get(column_name)
        if explanation and Comment is not None:
            cell.comment = Comment(explanation, "Coreg QC")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=False)

    numeric_columns = {
        "similarity_metric_value",
    }

    integer_columns = {
        "dof",
        "original_voxel_count",
        "used_voxel_count",
        "histogram_bins",
    }

    for col_idx, col_name in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)

        if col_name in numeric_columns:
            for cell in ws[col_letter][1:]:
                cell.number_format = "0.000000"

        elif col_name in integer_columns:
            for cell in ws[col_letter][1:]:
                cell.number_format = "0"

    last_col = get_column_letter(ws.max_column)
    last_row = max(ws.max_row, 1)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    for col_idx, col_name in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)

        max_length = len(str(col_name))

        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue

            value_str = str(value)
            first_line = max(value_str.splitlines() or [""], key=len)
            max_length = max(max_length, len(first_line))

        width = max_length + 2

        if col_name in {
            "session_dir",
            "moving_file",
            "reference_file",
            "output_file",
            "transform_file",
            "output_json",
            "metrics_error",
        }:
            width = min(max(width, 25), 85)
        elif col_name in numeric_columns:
            width = min(max(width, 14), 20)
        elif col_name in integer_columns:
            width = min(max(width, 10), 16)
        elif col_name in {"interpretation"}:
            width = min(max(width, 35), 70)
        else:
            width = min(max(width, 10), 32)

        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 36

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 18

    ws.sheet_view.showGridLines = True

    explanation_ws = wb.create_sheet("Rubrikförklaringar")
    explanation_ws.append(["Rubrik", "Förklaring"])

    for column in columns:
        explanation_ws.append([
            column,
            SIMILARITY_COLUMN_EXPLANATIONS.get(column, ""),
        ])

    for cell in explanation_ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in explanation_ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    explanation_ws.freeze_panes = "A2"
    explanation_ws.auto_filter.ref = f"A1:B{max(explanation_ws.max_row, 1)}"
    explanation_ws.column_dimensions["A"].width = 32
    explanation_ws.column_dimensions["B"].width = 90

    for row_idx in range(2, explanation_ws.max_row + 1):
        explanation_ws.row_dimensions[row_idx].height = 36

    wb.save(path)


def infer_subject_session_from_registration_json(
    root_dir: Path,
    json_path: Path,
) -> Tuple[str, str]:
    """
    Försöker läsa ut subject och session från sökvägen till en registration-JSON.

    Funktionen använder i första hand derivatives/registration-strukturen.
    Om det misslyckas söker den efter mappar som börjar med sub- och ses-.
    """

    try:
        relative = json_path.relative_to(root_dir / "derivatives" / "registration")
        parts = relative.parts
        if len(parts) >= 2 and parts[0].startswith("sub-") and parts[1].startswith("ses-"):
            return parts[0], parts[1]
    except Exception:
        pass

    subject = "unknown_subject"
    session = "unknown_session"

    for part in json_path.parts:
        if part.startswith("sub-"):
            subject = part
        if part.startswith("ses-"):
            session = part

    return subject, session


def collect_existing_similarity_rows_from_outputs(root_dir: Path) -> List[Dict[str, Any]]:
    """
    Skapar similarity metric-rader från redan befintliga coregistreringsfiler.

    Funktionen används när användaren vill skapa eller uppdatera Excel-filen
    utan att köra om själva coregistreringen.
    """

    summary_root = root_dir / "derivatives" / "registration"

    if not summary_root.exists():
        raise FileNotFoundError(
            f"Hittade ingen registration-mapp: {summary_root}"
        )

    rows: List[Dict[str, Any]] = []

    for json_path in sorted(summary_root.rglob("*_space-T1w*.json")):
        name_lower = json_path.name.lower()

        if "_desc-coreg" not in name_lower and "_space-t1w_coreg" not in name_lower:
            continue

        try:
            data = load_json_lenient(str(json_path))
            reg_block = data.get("Registration", {})

            if not isinstance(reg_block, dict):
                continue

            reference_file = data.get("ReferenceNIfTIFile") or reg_block.get("ReferenceFile")
            output_file = data.get("OutputNIfTIFile") or reg_block.get("OutputFile")
            moving_file = data.get("InputNIfTIFile")
            transform_file = reg_block.get("TransformFile")
            cost_function = reg_block.get("CostFunction")
            dof = reg_block.get("DegreesOfFreedom")

            if not reference_file or not output_file:
                continue

            if not Path(reference_file).exists() or not Path(output_file).exists():
                metrics = {
                    "computed": False,
                    "error": "ReferenceNIfTIFile eller OutputNIfTIFile saknas på disk.",
                }
            else:
                metrics = compute_similarity_metrics_safe(
                    reference_file=str(reference_file),
                    registered_file=str(output_file),
                )

            subject, session = infer_subject_session_from_registration_json(
                root_dir=root_dir,
                json_path=json_path,
            )

            session_dir = root_dir / subject / session

            moving_role = reg_block.get("MovingRole", "")
            moving_role_lower = str(moving_role).lower()

            if "pet" in moving_role_lower:
                coregistration = "pet"
            elif "t2" in moving_role_lower:
                coregistration = "t2"
            elif "dwi" in moving_role_lower or "diffusion" in moving_role_lower:
                coregistration = "diffusion"
            elif "perfusion" in moving_role_lower:
                coregistration = "perfusion"
            else:
                coregistration = "unknown"

            reg_for_row = {
                "status": "existing_output",
                "cost_function": cost_function,
                "dof": dof,
                "moving_file": moving_file,
                "reference_file": reference_file,
                "output_file": output_file,
                "transform_file": transform_file,
                "output_json": str(json_path),
                "similarity_metrics": metrics,
            }

            row = similarity_row_from_registration(
                subject=subject,
                session=session,
                session_dir=str(session_dir),
                coregistration=coregistration,
                reg=reg_for_row,
            )

            rows.append(row)

        except Exception:
            rows.append({
                "subject": "unknown_subject",
                "session": "unknown_session",
                "coregistration": "unknown",
                "status": "failed_reading_json",
                "cost_function": None,
                "similarity_metric_key": None,
                "similarity_metric_name": None,
                "similarity_metric_short_name": None,
                "similarity_metric_value": None,
                "better_direction": None,
                "interpretation": None,
                "dof": None,
                "metrics_computed": False,
                "metrics_error": traceback.format_exc(),
                "session_dir": None,
                "moving_file": None,
                "reference_file": None,
                "output_file": None,
                "transform_file": None,
                "output_json": str(json_path),
            })

    return rows


def open_file_with_default_app(path: Path) -> None:
    """
    Öppnar en fil med operativsystemets standardprogram.

    Använder os.startfile på Windows, open på macOS och xdg-open på Linux.
    """
    
    if not path.exists():
        raise FileNotFoundError(path)

    if os.name == "nt":
        os.startfile(str(path))  # type: ignore[attr-defined]
    elif sys.platform == "darwin":
        subprocess.Popen(["open", str(path)])
    else:
        subprocess.Popen(["xdg-open", str(path)])


def extract_common_json_fields_from_nifti(info: Dict[str, Any]) -> Dict[str, Any]:
    """
    Plockar ut NIfTI-metadata som ska sparas i JSON-sidecar-filen.

    Exempel på fält är shape, voxelstorlek, affine, datatyp och orientering.
    """

    voxel_sizes = info.get("voxel_sizes", [])
    out: Dict[str, Any] = {
        "NIfTIShape": info.get("shape"),
        "NIfTIVoxelSizes": voxel_sizes,
        "NIfTIAffine": info.get("affine"),
        "NIfTIAxisCodes": info.get("axis_codes"),
        "NIfTIDataType": info.get("dtype"),
    }

    if voxel_sizes:
        out["SliceThickness"] = float(voxel_sizes[2]) if len(voxel_sizes) >= 3 else float(voxel_sizes[-1])
        if len(voxel_sizes) >= 3:
            out["SpacingBetweenSlices"] = float(voxel_sizes[2])

    if "ImageOrientationPatientDICOM" in info:
        out["ImageOrientationPatientDICOM"] = info["ImageOrientationPatientDICOM"]

    return out


def matrix_file_to_list(path: str) -> List[List[float]]:
    """
    Läser en FLIRT-transformationsmatris från fil.

    Returnerar matrisen som en lista av listor så att den kan sparas i JSON.
    """

    rows: List[List[float]] = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            stripped = line.strip()
            if not stripped:
                continue
            rows.append([float(x) for x in stripped.split()])
    return rows


def basename_without_known_suffixes(path: Path) -> str:
    """
    Tar bort kända filändelser från ett filnamn.

    Hanterar särskilt .nii.gz, .nii och .json.
    """

    name = path.name
    if name.endswith(".nii.gz"):
        return name[:-7]
    if name.endswith(".nii"):
        return name[:-4]
    if name.endswith(".json"):
        return name[:-5]
    return path.stem


def bids_label(value: str) -> str:
    """
    Gör en BIDS-säker label: endast bokstäver och siffror.
    Exempel: cost-normmi -> CostNormmi
    """

    parts = re.split(r"[^A-Za-z0-9]+", value.strip())
    cleaned = "".join(part[:1].upper() + part[1:] for part in parts if part)
    return cleaned or "Coreg"


def split_bids_stem_prefix_and_suffix(path: Path) -> Tuple[str, str]:
    """
    Delar ett BIDS-liknande filnamn i prefix och suffix.

    Hanterar även interna filer som slutar med _desc-b0 eller _desc-ref.
    """

    stem = basename_without_known_suffixes(path)
    parts = stem.split("_")

    if len(parts) < 2:
        return stem, "image"

    # Om filen är en intern derived reference, t.ex.
    # sub-001_ses-01_dwi_desc-b0
    # ska suffix fortfarande bli dwi, inte desc-b0.
    if parts[-1].startswith("desc-") and len(parts) >= 3:
        suffix = parts[-2]
        prefix = "_".join(parts[:-2])
        return prefix, suffix

    prefix, suffix = stem.rsplit("_", 1)
    return prefix, suffix


def bids_coreg_derivative_stem(
    moving_path: Path,
    desc_label: str = "coreg",
) -> str:
    """
    Skapar ett BIDS-liknande filnamn för en coregistrerad derivative.

    Outputnamnet innehåller space-T1w för att markera att bilden ligger i
    T1-space, samt desc-coreg för att visa att den är skapad av registreringen.
    
    Exempel:
    sub-001_ses-01_pet.nii.gz
    ->
    sub-001_ses-01_space-T1w_desc-coreg_pet.nii.gz
    """

    prefix, suffix = split_bids_stem_prefix_and_suffix(moving_path)

    desc = bids_label(desc_label)
    desc = desc[:1].lower() + desc[1:]

    return f"{prefix}_space-T1w_desc-{desc}_{suffix}"


def companion_json_path(image_path: Path) -> Optional[Path]:
    """
    Letar efter en JSON-sidecar med samma basnamn som en bildfil.

    Returnerar sökvägen om JSON-filen finns, annars None.
    """

    base = basename_without_known_suffixes(image_path)
    json_path = image_path.with_name(base + ".json")
    return json_path if json_path.exists() else None


def companion_bval_path(image_path: Path) -> Optional[Path]:
    """
    Letar efter en .bval-fil som hör till en diffusion-NIfTI-fil.
    """

    base = basename_without_known_suffixes(image_path)
    bval_path = image_path.with_name(base + ".bval")
    return bval_path if bval_path.exists() else None


def companion_bvec_path(image_path: Path) -> Optional[Path]:
    """
    Letar efter en .bvec-fil som hör till en diffusion-NIfTI-fil.
    """

    base = basename_without_known_suffixes(image_path)
    bvec_path = image_path.with_name(base + ".bvec")
    return bvec_path if bvec_path.exists() else None


def diffusion_sidecar_paths(image_path: Path) -> Tuple[Path, Path]:
    """
    Skapar förväntade sökvägar till .bval och .bvec för en diffusionfil.
    """

    base = basename_without_known_suffixes(image_path)
    return (
        image_path.with_name(base + ".bval"),
        image_path.with_name(base + ".bvec"),
    )


def write_single_b0_bval_bvec(output_nifti: Path) -> None:
    """
    Skapar BIDS-required .bval och .bvec för en 3D B0-DWI derivative.
    """

    out_bval, out_bvec = diffusion_sidecar_paths(output_nifti)

    out_bval.write_text("0\n", encoding="utf-8")
    out_bvec.write_text("0\n0\n0\n", encoding="utf-8")


def ensure_dwi_bval_bvec_for_registration_outputs(registration_root: Path) -> int:
    """
    Skapar saknade .bval och .bvec för redan skapade coreg-DWI-filer.
    """

    updated_count = 0

    for nifti_path in sorted(registration_root.rglob("*_space-T1w*_dwi.nii.gz")):
        if not is_coregistered_to_t1_file(nifti_path):
            continue

        bval_path, bvec_path = diffusion_sidecar_paths(nifti_path)

        changed = False

        if not bval_path.exists():
            bval_path.write_text("0\n", encoding="utf-8")
            changed = True

        if not bvec_path.exists():
            bvec_path.write_text("0\n0\n0\n", encoding="utf-8")
            changed = True

        if changed:
            updated_count += 1

    return updated_count


def extract_b0_volume(
    dwi_path: Path,
    out_dir: Path,
    b0_threshold: float = 75.0,
) -> Path:
    """
    Extraherar första B0-volymen från en 4D diffusion-fil.

    B0-volymen används som representativ 3D-bild för att kunna registrera
    diffusionen mot T1. Volymer med b-värde under tröskeln betraktas som B0.
    """

    if nib is None or np is None:
        raise RuntimeError("B0-extraktion kräver nibabel och numpy.")

    bval_path = companion_bval_path(dwi_path)
    if bval_path is None:
        raise FileNotFoundError(f"Hittade ingen .bval till diffusion-filen: {dwi_path}")

    img = nib.load(str(dwi_path))
    data = np.asanyarray(img.dataobj)

    if data.ndim != 4:
        raise RuntimeError(f"Diffusion-filen är inte 4D: {dwi_path}")

    bvals = np.loadtxt(str(bval_path))
    bvals = np.atleast_1d(bvals).astype(float)

    if data.shape[3] != len(bvals):
        raise RuntimeError(
            f"Antal volymer i DWI ({data.shape[3]}) matchar inte antal b-värden ({len(bvals)}): {dwi_path}"
        )

    b0_indices = np.where(bvals <= b0_threshold)[0]
    if len(b0_indices) == 0:
        raise RuntimeError(f"Hittade ingen B0-volym i {bval_path}")

    b0_index = int(b0_indices[0])

    out_dir.mkdir(parents=True, exist_ok=True)

    base = basename_without_known_suffixes(dwi_path)
    out_b0 = out_dir / f"{base}_desc-b0.nii.gz"
    out_json = out_dir / f"{base}_desc-b0.json"

    b0_data = data[..., b0_index]
    b0_img = nib.Nifti1Image(b0_data, img.affine, img.header.copy())
    nib.save(b0_img, str(out_b0))

    src_json_path = companion_json_path(dwi_path)
    src_json = load_json_lenient(str(src_json_path)) if src_json_path else {}

    b0_json = dict(src_json)
    b0_json["SourceDWIFile"] = str(dwi_path)
    b0_json["SourceBvalFile"] = str(bval_path)
    b0_json["B0VolumeIndex"] = b0_index
    b0_json["B0Threshold"] = float(b0_threshold)
    b0_json["DerivedFrom"] = str(dwi_path)

    save_json(str(out_json), b0_json)

    return out_b0


def extract_perfusion_reference_volume(
    perf_path: Path,
    out_dir: Path,
    method: str = "mean",
) -> Path:
    """
    Skapar en 3D-referensvolym från perfusionsdata.

    Om perfusionsfilen är 4D används medelvärdet över tid som referensbild.
    Detta gör att en stabil 3D-bild kan registreras mot T1.
    """

    if nib is None or np is None:
        raise RuntimeError("Perfusion-extraktion kräver nibabel och numpy.")

    img = nib.load(str(perf_path))
    data = np.asanyarray(img.dataobj)

    if data.ndim == 3:
        return perf_path

    if data.ndim != 4:
        raise RuntimeError(f"Perfusionsfilen är varken 3D eller 4D: {perf_path}")

    out_dir.mkdir(parents=True, exist_ok=True)

    base = basename_without_known_suffixes(perf_path)
    out_img_path = out_dir / f"{base}_desc-ref.nii.gz"
    out_json_path = out_dir / f"{base}_desc-ref.json"

    if method == "mean":
        ref_data = np.mean(data, axis=3)
        ref_desc = "mean_over_time"
    else:
        ref_data = data[..., 0]
        ref_desc = "first_volume"

    ref_header = img.header.copy()
    ref_img = nib.Nifti1Image(ref_data, img.affine, ref_header)
    nib.save(ref_img, str(out_img_path))

    src_json_path = companion_json_path(perf_path)
    src_json = load_json_lenient(str(src_json_path)) if src_json_path else {}

    out_json = dict(src_json)
    out_json["DerivedFrom"] = str(perf_path)
    out_json["SourcePerfusionFile"] = str(perf_path)
    out_json["ReferenceVolumeMethod"] = ref_desc

    save_json(str(out_json_path), out_json)

    return out_img_path


def is_generated_file(path: Path) -> bool:
    """
    Kontrollerar om en fil verkar vara skapad av programmet.

    Detta används för att undvika att tidigare outputfiler används som ny input.
    """

    lower = path.name.lower()
    return any(marker in lower for marker in GENERATED_MARKERS)


def is_coregistered_to_t1_file(path: Path) -> bool:
    """
    Kontrollerar om en NIfTI-fil verkar vara coregistrerad till T1-space.

    Funktionen känner igen både äldre och nyare filnamnsformat.
    """

    lower = path.name.lower()

    if not is_nifti_file(str(path)):
        return False

    if is_checkerboard_qc_file(path):
        return False

    # Gammalt namn:
    # *_space-T1w_coreg*.nii.gz
    if "_space-t1w_coreg" in lower:
        return True

    # Nytt BIDS-liknande namn:
    # *_space-T1w_desc-coreg*_T2w.nii.gz
    if "_space-t1w" in lower and "_desc-coreg" in lower:
        return True

    return False


def find_matching_nifti_files(session_dir: Path, token: str) -> List[Path]:
    """
    Söker efter NIfTI-filer i en session vars filnamn innehåller ett visst token.

    Filer som programmet själv har skapat ignoreras.
    """

    matches: List[Path] = []
    token = token.lower()

    for path in session_dir.rglob("*"):
        if not path.is_file():
            continue
        if not is_nifti_file(str(path)):
            continue
        if is_generated_file(path):
            continue
        if token in path.name.lower():
            matches.append(path)

    return sorted(matches)


def choose_unique_match(session_dir: Path, token: str, label: str, required: bool) -> Optional[Path]:
    """
    Väljer exakt en matchande NIfTI-fil.

    Om ingen fil hittas och required är True kastas ett fel.
    Om flera filer hittas kastas också ett fel eftersom valet är tvetydigt.
    """

    matches = find_matching_nifti_files(session_dir, token)

    if not matches:
        if required:
            raise FileNotFoundError(
                f"Hittade ingen {label}-fil i {session_dir} som matchar '{token}'."
            )
        return None

    if len(matches) > 1:
        pretty = "\n".join(f"  - {m}" for m in matches)
        raise RuntimeError(
            f"Hittade flera {label}-filer i {session_dir} som matchar '{token}'.\n{pretty}"
        )

    return matches[0]


def find_subject_session_dirs(root_dir: Path) -> List[Tuple[str, str, Path]]:
    """
    Söker igenom rotmappen och hittar alla sessioner som följer strukturen
    sub-*/ses-*.
    """

    found: List[Tuple[str, str, Path]] = []

    for sub_dir in sorted(root_dir.glob("sub-*")):
        if not sub_dir.is_dir():
            continue
        for ses_dir in sorted(sub_dir.glob("ses-*")):
            if ses_dir.is_dir():
                found.append((sub_dir.name, ses_dir.name, ses_dir))

    return found


def list_nifti_with_json(folder: Path) -> List[Path]:
    """
    Returnerar alla NIfTI-filer i en mapp som också har en JSON-sidecar.

    Detta gör att programmet främst arbetar med filer som har tillhörande metadata.
    """

    if not folder.exists():
        return []

    out: List[Path] = []
    for path in folder.rglob("*"):
        if not path.is_file():
            continue
        if not is_nifti_file(str(path)):
            continue
        if companion_json_path(path) is None:
            continue
        out.append(path)
    return sorted(out)


def choose_best_candidate(candidates: List[Path], label: str, scorer) -> Optional[Path]:
    """
    Väljer bästa kandidat från en lista med filer.

    scorer-funktionen bestämmer prioriteringen. Lägst score väljs.
    """

    if not candidates:
        return None

    ranked = sorted(candidates, key=scorer)
    best = ranked[0]
    return best


def find_t1_file(session_dir: Path, use_contrast: bool) -> Optional[Path]:
    """
    Letar upp bästa T1-kandidat i en session.

    Om use_contrast är True prioriteras kontrastförstärkta T1-bilder.
    Annars prioriteras vanliga T1w-bilder utan kontrast.
    """

    anat_dir = session_dir / "anat"
    candidates = [
        p for p in list_nifti_with_json(anat_dir)
        if "_t1w" in p.name.lower()
    ]

    def score(path: Path) -> Tuple[int, str]:
        name = path.name.lower()

        if use_contrast:
            if name.endswith("_ce-gd_t1w.nii.gz") or name.endswith("_ce-gd_t1w.nii"):
                return (0, name)
            if "_acq-cubefs_ce-gd_t1w" in name:
                return (1, name)
            if "ce-gd" in name and "_t1w" in name:
                return (2, name)
            return (100, name)

        if name.endswith("_t1w.nii.gz") or name.endswith("_t1w.nii"):
            if "ce-gd" not in name and "_acq-" not in name:
                return (0, name)
        if "_acq-cube_t1w" in name and "ce-gd" not in name:
            return (1, name)
        if "_acq-fmri_t1w" in name:
            return (50, name)
        if "ce-gd" in name:
            return (100, name)
        return (10, name)

    return choose_best_candidate(candidates, "T1", score)


def find_t2_file(session_dir: Path, use_contrast: bool) -> Optional[Path]:
    """
    Letar upp T2-filen i sessionens anat-mapp.
    """

    anat_dir = session_dir / "anat"
    candidates = [
        p for p in list_nifti_with_json(anat_dir)
        if "_t2w" in p.name.lower()
    ]

    def score(path: Path) -> Tuple[int, str]:
        name = path.name.lower()

        if use_contrast:
            if name.endswith("_ce-gd_t2w.nii.gz") or name.endswith("_ce-gd_t2w.nii"):
                return (0, name)
            if "ce-gd" in name and "_t2w" in name:
                return (1, name)
            return (100, name)

        if name.endswith("_t2w.nii.gz") or name.endswith("_t2w.nii"):
            if "ce-gd" not in name:
                return (0, name)
        if "ce-gd" in name:
            return (100, name)
        return (10, name)

    return choose_best_candidate(candidates, "T2", score)


def find_pet_file(session_dir: Path) -> Optional[Path]:
    """
    Letar upp PET-filen i sessionens pet-mapp.

    Om flera PET-filer finns prioriteras rekonstruktioner med macstatic3mm
    och därefter macstatic.
    """

    pet_dir = session_dir / "pet"
    candidates = [
        p for p in list_nifti_with_json(pet_dir)
        if p.name.lower().endswith("_pet.nii.gz") or p.name.lower().endswith("_pet.nii")
    ]

    def score(path: Path) -> Tuple[int, str]:
        name = path.name.lower()

        if "_rec-macstatic3mm_pet" in name:
            return (0, name)

        if "_rec-macstatic_pet" in name:
            return (1, name)

        return (10, name)

    return choose_best_candidate(candidates, "PET", score)


def find_diffusion_file(session_dir: Path) -> Optional[Path]:
    """
    Letar upp Diffusion-filen i sessionens dwi-mapp.
    """

    dwi_dir = session_dir / "dwi"
    candidates = [
        p for p in list_nifti_with_json(dwi_dir)
        if "_dwi" in p.name.lower()
    ]
    return sorted(candidates)[0] if candidates else None


def find_perfusion_file(session_dir: Path) -> Optional[Path]:
    """
    Letar upp Perfusion-filen i sessionens perf-mapp.
    """

    perf_dir = session_dir / "perf"

    candidates = []
    for p in list_nifti_with_json(perf_dir):
        name = p.name.lower()

        if "_acq-dsc_perf" in name:
            candidates.append((0, p))
        elif name.endswith("_asl.nii.gz") or name.endswith("_asl.nii"):
            candidates.append((1, p))

    if not candidates:
        return None

    return sorted(candidates, key=lambda x: (x[0], str(x[1])))[0][1]


def registration_dir_for_session(root_dir: Path, subject: str, session: str) -> Path:
    """
    Returnerar outputmappen för registration-derivatives för en viss session.
    """

    return root_dir / "derivatives" / "registration" / subject / session


def path_for_generated_output(
    root_dir: Path,
    session_dir: Path,
    subject: str,
    session: str,
    moving_path: Path,
    output_subdir: Optional[Path] = None,
    name_suffix: str = "",
) -> Tuple[Path, Path, Path, Path]:
    """
    Skapar standardiserade output-sökvägar för en coregistrering.

    Returnerar sökvägar till output-NIfTI, JSON-sidecar, FLIRT-matris och
    kommandologg.
    """
    
    reg_session_dir = registration_dir_for_session(root_dir, subject, session)

    if output_subdir is not None:
        relative_parent = output_subdir
    else:
        try:
            relative_parent = moving_path.relative_to(session_dir).parent
        except ValueError:
            relative_parent = Path(moving_path.parent.name)

    out_dir = reg_session_dir / relative_parent
    out_dir.mkdir(parents=True, exist_ok=True)

    if name_suffix:
        desc_label = f"coreg-{name_suffix}"
    else:
        desc_label = "coreg"

    out_stem = bids_coreg_derivative_stem(
        moving_path=moving_path,
        desc_label=desc_label,
    )

    out_nifti = out_dir / f"{out_stem}.nii.gz"
    out_json = out_dir / f"{out_stem}.json"

    # .mat och command-loggar är inte BIDS-filer, så de ignoreras via .bidsignore.
    out_mat = out_dir / f"{out_stem}.mat"
    out_log = out_dir / f"{out_stem}_commands.txt"

    return out_nifti, out_json, out_mat, out_log

def find_coreg_files_for_session(root_dir: Path, session_dir: Path) -> List[Path]:
    """
    Letar upp färdiga coregistrerade filer för en session.

    Programmet söker först i derivatives/registration. Om inga filer hittas
    görs en fallback-sökning i själva sessionens mapp.
    """

    subject = session_dir.parent.name
    session = session_dir.name

    candidates: List[Path] = []

    reg_dir = registration_dir_for_session(root_dir, subject, session)
    if reg_dir.exists():
        for path in reg_dir.rglob("*"):
            if not path.is_file():
                continue
            if not is_nifti_file(str(path)):
                continue
            if is_checkerboard_qc_file(path):
                continue
            if not is_coregistered_to_t1_file(path):
                continue
            candidates.append(path)

    if not candidates:
        for path in session_dir.rglob("*"):
            if not path.is_file():
                continue
            if not is_nifti_file(str(path)):
                continue
            if is_checkerboard_qc_file(path):
                continue
            if not is_coregistered_to_t1_file(path):
                continue
            candidates.append(path)

    return sorted(candidates)

def find_reference_t1_for_session(root_dir: Path, session_dir: Path) -> Optional[Path]:
    """
    Försöker hitta T1-referensen som användes för en session.

    I första hand läses referensen från registration-JSON-filer.
    Om det inte går söker funktionen efter T1-filer i originalsessionen.
    """

    subject = session_dir.parent.name
    session = session_dir.name
    reg_dir = registration_dir_for_session(root_dir, subject, session)

    if reg_dir.exists():
        for json_path in sorted(reg_dir.rglob("*_space-T1w*.json")):
            if "_desc-coreg" not in json_path.name.lower() and "_space-t1w_coreg" not in json_path.name.lower():
                continue
            try:
                data = load_json_lenient(str(json_path))
                ref = data.get("ReferenceNIfTIFile")
                if ref and Path(ref).exists():
                    return Path(ref)
            except Exception:
                pass

    t1_matches = find_matching_nifti_files(session_dir, T1_TOKEN)
    if t1_matches:
        return t1_matches[0]

    anat_dir = session_dir / "anat"
    if anat_dir.exists():
        all_t1: List[Path] = []
        non_contrast_t1: List[Path] = []

        for path in anat_dir.rglob("*"):
            if not path.is_file():
                continue
            if not is_nifti_file(str(path)):
                continue
            lower = path.name.lower()
            if "_t1w" not in lower:
                continue
            all_t1.append(path)
            if "ce-gd" not in lower:
                non_contrast_t1.append(path)

        if non_contrast_t1:
            return sorted(non_contrast_t1)[0]
        if all_t1:
            return sorted(all_t1)[0]

    return None


def session_view_priority(path: Path) -> Tuple[int, str]:
    """
    Bestämmer sorteringsordning när filer öppnas i FSL.

    T1-referensen prioriteras först, därefter coregistrerade filer.
    """

    name = path.name.lower()

    if not is_coregistered_to_t1_file(path) and "_t1w" in name:
        return (0, name)

    if is_coregistered_to_t1_file(path):
        return (1, name)

    return (2, name)


def registration_json_block(
    moving_role: str,
    reference_role: str,
    matrix_path: str,
    output_path: str,
    dof: int,
    cost: str,
    flirt_meta: Dict[str, Any],
) -> Dict[str, Any]:
    """
    Skapar metadata-blocket som beskriver en coregistrering.

    Blocket innehåller information om inputroll, referensroll, FLIRT-metod,
    cost-function, degrees of freedom, kommandon och transformationsmatris.
    """

    return {
        "Performed": True,
        "MovingRole": moving_role,
        "ReferenceRole": reference_role,
        "OutputFile": str(output_path),
        "TransformFile": str(matrix_path),
        "Method": "Nipype FSL FLIRT + ApplyXFM",
        "CostFunction": cost,
        "DegreesOfFreedom": int(dof),
        "Interpolator": "trilinear",
        "SearchAngles": {
            "x": [-180, 180],
            "y": [-180, 180],
            "z": [-180, 180],
        },
        "Commands": {
            "Estimate": flirt_meta.get("estimate_command"),
            "Apply": flirt_meta.get("apply_command"),
        },
        "Matrix": flirt_meta.get("matrix"),
    }


def changed_image_fields(before: Dict[str, Any], after: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
    """
    Jämför metadata före och efter registrering.

    Endast fält som har ändrats tas med i resultatet. Detta används i
    ChangeLog-delen av output-JSON.
    """

    result: Dict[str, Dict[str, Any]] = {}
    for key in (
        "shape",
        "voxel_sizes",
        "dtype",
        "affine",
        "axis_codes",
        "ImageOrientationPatientDICOM",
        "data_min",
        "data_max",
    ):
        before_value = before.get(key)
        after_value = after.get(key)
        if before_value != after_value:
            result[key] = {
                "before": normalize_value_for_json(before_value),
                "after": normalize_value_for_json(after_value),
            }
    return result


def build_output_json(
    source_json: Optional[Dict[str, Any]],
    input_info: Dict[str, Any],
    output_info: Dict[str, Any],
    modality_kind: str,
    input_nifti_path: str,
    output_nifti_path: str,
    reference_nifti_path: str,
    input_json_path: Optional[str],
    output_json_path: str,
    registration_meta: Dict[str, Any],
) -> Dict[str, Any]:
    """
    Bygger den slutliga JSON-sidecaren för en coregistrerad outputfil.

    Funktionen kombinerar originalmetadata, ny NIfTI-metadata,
    BIDS-derivativefält, registreringsmetadata och en ChangeLog.
    """

    data: Dict[str, Any] = dict(source_json) if source_json else {}

    common = extract_common_json_fields_from_nifti(output_info)
    for k, v in common.items():
        data[k] = normalize_value_for_json(v)

    data["ProgramName"] = APP_TITLE
    data["InputNIfTIFile"] = str(input_nifti_path)
    data["InputJSONFile"] = str(input_json_path) if input_json_path else None
    data["ReferenceNIfTIFile"] = str(reference_nifti_path)
    # Krävs av BIDS Derivatives när output-filen har space-T1w i filnamnet.
    # Detta beskriver vilken bild outputen är alignad till.
    data["SpatialReference"] = str(reference_nifti_path)
    # Krävs av BIDS Derivatives för vissa MR-derivatives.
    # False betyder att vi inte explicit har skull-strippat bilden.
    data["SkullStripped"] = False
    data["OutputNIfTIFile"] = str(output_nifti_path)
    data["OutputJSONFile"] = str(output_json_path)
    data["StandardizedOutputKind"] = modality_kind

    if modality_kind == "PET":
        data["Modality"] = "PT"
    else:
        data["Modality"] = "MR"

    data["Registration"] = registration_meta
    data["ChangeLog"] = {
        "CreatedAt": datetime.now().isoformat(timespec="seconds"),
        "InputImageInfo": input_info,
        "OutputImageInfo": output_info,
        "ChangedImageFields": changed_image_fields(input_info, output_info),
    }

    return data


def write_command_log(
    log_file: str,
    moving_file: str,
    reference_file: str,
    output_file: str,
    matrix_file: str,
    cost: str,
    dof: int,
    estimate_command: Optional[str],
    apply_command: Optional[str],
) -> None:
    """
    Skriver en textlogg med de FSL-kommandon som användes.

    Loggen gör det möjligt att i efterhand se exakt hur registreringen kördes.
    """

    lines = [
        f"Program: {APP_TITLE}",
        f"CreatedAt: {datetime.now().isoformat(timespec='seconds')}",
        "",
        f"MovingFile: {moving_file}",
        f"ReferenceFile: {reference_file}",
        f"OutputFile: {output_file}",
        f"MatrixFile: {matrix_file}",
        f"CostFunction: {cost}",
        f"DegreesOfFreedom: {dof}",
        "",
        "EstimateCommand:",
        estimate_command or "N/A",
        "",
        "ApplyCommand:",
        apply_command or "N/A",
        "",
    ]
    with open(log_file, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


class UserCancelledError(RuntimeError):
    """
    Eget fel som används när användaren avbryter en pågående körning.
    """

    pass


def run_tracked_command(command: str, work_dir: str, runtime: RuntimeControl) -> str:
    """
    Kör ett externt kommando och håller reda på processen.

    Funktionen används främst för FSL-kommandon. Den stödjer avbrott,
    timeout, loggning av output och felhantering om kommandot misslyckas.
    """

    if runtime.cancel_event.is_set():
        raise UserCancelledError("Körningen avbröts av användaren.")

    env = os.environ.copy()
    env["FSLOUTPUTTYPE"] = "NIFTI_GZ"

    popen_kwargs = {
        "cwd": work_dir,
        "env": env,
        "stdout": subprocess.PIPE,
        "stderr": subprocess.STDOUT,
        "text": True,
    }

    if os.name != "nt":
        popen_kwargs["start_new_session"] = True

    proc = subprocess.Popen(shlex.split(command), **popen_kwargs)
    runtime.register(proc)

    try:
        try:
            output, _ = proc.communicate(timeout=FSL_COMMAND_TIMEOUT_SECONDS)
        except subprocess.TimeoutExpired:
            if os.name != "nt":
                os.killpg(proc.pid, signal.SIGTERM)
            else:
                proc.terminate()

            try:
                output, _ = proc.communicate(timeout=20)
            except subprocess.TimeoutExpired:
                if os.name != "nt":
                    os.killpg(proc.pid, signal.SIGKILL)
                else:
                    proc.kill()
                output, _ = proc.communicate()

            raise RuntimeError(
                "FSL-kommandot tog för lång tid och stoppades.\n\n"
                f"Timeout: {FSL_COMMAND_TIMEOUT_SECONDS} sekunder\n"
                f"Kommando:\n{command}\n\n"
                f"Output:\n{output or ''}"
            )
    finally:
        runtime.unregister(proc)

    if runtime.cancel_event.is_set():
        raise UserCancelledError("Körningen avbröts av användaren.")

    if proc.returncode != 0:
        raise RuntimeError(
            f"Kommandot misslyckades med exitkod {proc.returncode}:\n"
            f"{command}\n\n{output or ''}"
        )

    return output or ""

###____________Works in terminal______________________________________
# def build_bids_validator_command(
#     root_dir: Path,
#     prune_derivatives: bool = True,
# ) -> List[str]:
#     """
#     Bygger kommandot för BIDS Validator.

#     Prioritet:
#     1. Använd installerad bids-validator om den finns i PATH.
#     2. Annars använd Deno + jsr:@bids/validator.
#     """
#     installed_validator = find_executable("bids-validator")

#     if installed_validator is not None:
#         cmd = [installed_validator]
#     else:
#         deno = find_executable("deno")

#         if deno is None:
#             deno_candidates = [
#                 Path.home() / ".deno" / "bin" / "deno",
#                 Path("/opt/anaconda3/envs/BIDS_and_coreg/bin/deno"),
#             ]

#             for deno_candidate in deno_candidates:
#                 if is_executable_file(deno_candidate):
#                     deno = str(deno_candidate)
#                     break

#         if deno is None:
#             raise RuntimeError(
#                 "BIDS Validator kunde inte hittas.\n\n"
#                 "Installera Deno, till exempel med:\n"
#                 "  conda install -c conda-forge deno\n\n"
#                 "Testa sedan i terminalen:\n"
#                 "  deno run -ERWN jsr:@bids/validator --prune --ignoreWarnings /path/to/bids_dataset"
#             )

#         ### OBS: Manfred ändrar
#         if getattr(sys, 'frozen', False):
#             bundle_dir = Path(sys._MEIPASS)
#         else:
#             bundle_dir = Path(__file__).parent

#         deno_path = bundle_dir / "deno"

#         cmd = [
#             deno_path,
#             "run",
#             "-ERWN",
#             "jsr:@bids/validator",
#         ]

#     if prune_derivatives:
#         cmd.append("--prune")

#     # Viktigt: stoppa inte batchen för rekommenderade metadata-varningar.
#     cmd.append("--ignoreWarnings")

#     # Tar bort ANSI-färgkoder från outputen i Tkinter-rutan.
#     cmd.append("--no-color")

#     cmd.append(str(root_dir))

#     return cmd
###_______________________________________________________________________

###___Works in app________________________________________________________
def build_bids_validator_command(
    root_dir: Path,
    prune_derivatives: bool = True,
) -> List[str]:
    
    if getattr(sys, "frozen", False):
        bundle_dir = Path(sys.executable).parent
    else:
        bundle_dir = Path(__file__).parent

    # Bundled deno first
    deno_candidates = [
        bundle_dir / "deno",
        Path("/opt/anaconda3/envs/BIDS_and_coreg/bin/deno"),
        Path.home() / ".deno" / "bin" / "deno",
    ]

    deno = None

    for candidate in deno_candidates:
        if candidate.exists() and os.access(candidate, os.X_OK):
            deno = str(candidate)
            break

    if deno is None:
        raise RuntimeError(
            "Deno kunde inte hittas."
        )

    cmd = [
        deno,
        "run",
        "-ERWN",
        "jsr:@bids/validator",
    ]

    if prune_derivatives:
        cmd.append("--prune")

    cmd.append("--ignoreWarnings")
    cmd.append("--no-color")
    cmd.append(str(root_dir))

    return cmd
###________________________________________________________________________

def run_bids_validator(
    root_dir: Path,
    runtime: Optional[RuntimeControl] = None,
    prune_derivatives: bool = True,
) -> Dict[str, Any]:
    """
    Kör BIDS Validator på rotmappen.

    Returnerar dict med:
    - valid: True/False
    - returncode
    - command
    - output
    """
    if runtime is not None and runtime.cancel_event.is_set():
        raise UserCancelledError("Körningen avbröts av användaren.")

    cmd = build_bids_validator_command(
        root_dir=root_dir,
        prune_derivatives=prune_derivatives,
    )

    popen_kwargs = {
        "cwd": str(root_dir),
        "stdout": subprocess.PIPE,
        "stderr": subprocess.STDOUT,
        "text": True,
    }

    if os.name != "nt":
        popen_kwargs["start_new_session"] = True

    proc = subprocess.Popen(cmd, **popen_kwargs)

    if runtime is not None:
        runtime.register(proc)

    try:
        output, _ = proc.communicate()
    finally:
        if runtime is not None:
            runtime.unregister(proc)

    if runtime is not None and runtime.cancel_event.is_set():
        raise UserCancelledError("BIDS-valideringen avbröts av användaren.")

    return {
        "valid": proc.returncode == 0,
        "returncode": proc.returncode,
        "command": " ".join(shlex.quote(part) for part in cmd),
        "output": output or "",
    }


def write_bids_validator_report(
    path: Path,
    root_dir: Path,
    validation_result: Dict[str, Any],
) -> None:
    """
    Skriver en textbaserad rapport från BIDS Validator.

    Rapporten innehåller kommando, return code, valid-status och validator-output.
    """

    lines = [
        "BIDS Validator report",
        f"CreatedAt: {datetime.now().isoformat(timespec='seconds')}",
        f"RootDirectory: {root_dir}",
        f"Command: {validation_result.get('command')}",
        f"ReturnCode: {validation_result.get('returncode')}",
        f"Valid: {validation_result.get('valid')}",
        "",
        "Output:",
        "",
        validation_result.get("output") or "",
    ]

    path.parent.mkdir(parents=True, exist_ok=True)

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def ensure_derivative_dataset_description(registration_root: Path) -> None:
    """
    Säkerställer att derivatives/registration har de filer som behövs för BIDS.

    Funktionen skapar en minimal dataset_description.json och uppdaterar
    .bidsignore så att interna arbetsfiler, loggar och temporära derivatives
    inte orsakar fel i BIDS Validator.
    """

    dataset_description_path = registration_root / "dataset_description.json"

    if not dataset_description_path.exists():
        data = {
            "Name": "Registration derivatives",
            "BIDSVersion": "1.10.0",
            "DatasetType": "derivative",
            "GeneratedBy": [
                {
                    "Name": "registration",
                    "Description": APP_TITLE,
                }
            ],
        }

        save_json(str(dataset_description_path), data)

    bidsignore_path = registration_root / ".bidsignore"

    required_ignore_lines = [
        "/work/",
        "/work/**",
        "work/",
        "work/**",
        "**/work/",
        "**/work/**",

        "/checkerboard_qc/",
        "/checkerboard_qc/**",
        "checkerboard_qc/",
        "checkerboard_qc/**",

        "bids_validator_report.txt",
        "coregistration_batch_summary.json",
        "coregistration_similarity_metrics.xlsx",

        "**/*_space-T1w_coreg*",
        "**/*_space-t1w_coreg*",
        "*_space-T1w_coreg*",
        "*_space-t1w_coreg*",

        "*.txt",
        "*.xlsx",
        "*.mat",
        "*_commands.txt",

        "*_estimate_prelim.nii.gz",
        "*_desc-b0.nii.gz",
        "*_desc-b0.json",
        "*_desc-ref.nii.gz",
        "*_desc-ref.json",

        # Coreg-DWI derivatives ignoreras av BIDS Validator.
        # De är en 3D B0-bild i T1-space, inte en komplett raw DWI-serie.
        "**/*_space-T1w*_dwi.nii.gz",
        "**/*_space-T1w*_dwi.json",
        "**/*_space-T1w*_dwi.bval",
        "**/*_space-T1w*_dwi.bvec",
        "**/*_space-t1w*_dwi.nii.gz",
        "**/*_space-t1w*_dwi.json",
        "**/*_space-t1w*_dwi.bval",
        "**/*_space-t1w*_dwi.bvec",

        # Coreg-perfusion derivatives ignoreras av BIDS Validator.
        # De är interna registration outputs och matchar inte alltid en BIDS-validatorregel.
        "**/*_space-T1w*_perf.nii.gz",
        "**/*_space-T1w*_perf.json",
        "**/*_space-t1w*_perf.nii.gz",
        "**/*_space-t1w*_perf.json",

        # Coreg-PET/T2 derivatives ignoreras också för säkerhet.
        "**/*_space-T1w*_pet.nii.gz",
        "**/*_space-T1w*_pet.json",
        "**/*_space-t1w*_pet.nii.gz",
        "**/*_space-t1w*_pet.json",

        # Coreg-ASL derivatives ignoreras av BIDS Validator.
        # ASL kräver metadata som RepetitionTimePreparation, men dessa filer är interna registration outputs.
        "**/*_space-T1w*_asl.nii.gz",
        "**/*_space-T1w*_asl.json",
        "**/*_space-t1w*_asl.nii.gz",
        "**/*_space-t1w*_asl.json",

        "**/*_space-T1w*_T2w.nii.gz",
        "**/*_space-T1w*_T2w.json",
        "**/*_space-t1w*_t2w.nii.gz",
        "**/*_space-t1w*_t2w.json",

        # Robust catch-all: ignorera alla coregistrerade outputs.
        "**/*_space-T1w_desc-coreg*",
        "**/*_space-t1w_desc-coreg*",
        "**/*_space-T1w*_desc-coreg*",
        "**/*_space-t1w*_desc-coreg*",
    ]

    existing_lines: List[str] = []

    if bidsignore_path.exists():
        existing_text = bidsignore_path.read_text(encoding="utf-8")
        existing_lines = [
            line.strip()
            for line in existing_text.splitlines()
            if line.strip()
        ]

    merged_lines = list(existing_lines)

    for line in required_ignore_lines:
        if line not in merged_lines:
            merged_lines.append(line)

    bidsignore_path.write_text(
        "\n".join(merged_lines) + "\n",
        encoding="utf-8",
    )


def ensure_required_derivative_sidecar_keys(registration_root: Path) -> int:
    """
    Patchar redan skapade coreg-JSON-filer så att BIDS Validator
    inte stoppar på saknade derivative-metadatafält.
    """

    updated_count = 0

    for json_path in sorted(registration_root.rglob("*_space-T1w*.json")):
        name_lower = json_path.name.lower()

        if "_desc-coreg" not in name_lower and "_space-t1w_coreg" not in name_lower:
            continue

        try:
            data = load_json_lenient(str(json_path))
            changed = False

            reg_block = data.get("Registration", {})
            if not isinstance(reg_block, dict):
                reg_block = {}

            reference_file = (
                data.get("ReferenceNIfTIFile")
                or reg_block.get("ReferenceFile")
                or "T1w"
            )

            if not data.get("SpatialReference"):
                data["SpatialReference"] = str(reference_file)
                changed = True

            if "SkullStripped" not in data:
                data["SkullStripped"] = False
                changed = True

            if changed:
                save_json(str(json_path), data)
                updated_count += 1

        except Exception:
            pass

    return updated_count


def remove_registration_work_before_bids_validation(registration_root: Path) -> None:
    """
    Tar bort gammal intern work-mapp innan BIDS Validator körs.

    work/ innehåller temporära filer från FSL/Nipype och är inte BIDS-data.
    Slutresultaten ligger i sub-*/ses-* och påverkas inte.
    """

    work_dir = registration_root / "work"

    if work_dir.exists():
        if not work_dir.is_dir():
            raise RuntimeError(f"Förväntade att work var en mapp, men hittade fil: {work_dir}")

        shutil.rmtree(work_dir)


def remove_coreg_dwi_gradient_sidecars(registration_root: Path) -> int:
    """
    Tar bort .bval/.bvec som tidigare skapats för coreg-DWI derivatives.
    Dessa gör att BIDS Validator får filename-rule-problem.
    """

    removed_count = 0

    patterns = [
        "*_space-T1w*_dwi.bval",
        "*_space-T1w*_dwi.bvec",
        "*_space-t1w*_dwi.bval",
        "*_space-t1w*_dwi.bvec",
    ]

    for pattern in patterns:
        for path in registration_root.rglob(pattern):
            if path.is_file():
                path.unlink()
                removed_count += 1

    return removed_count


def estimate_flirt_transform(
    moving_file: str,
    reference_file: str,
    matrix_file: str,
    dof: int,
    cost: str,
    work_dir: str,
    runtime: RuntimeControl,
) -> Dict[str, Any]:
    """
    Estimerar transformationsmatrisen mellan moving-bilden och T1-referensen.

    Funktionen använder FSL FLIRT via Nipype. Resultatet blir en .mat-fil som
    beskriver hur moving-bilden ska transformeras till T1-space.
    """

    work_dir_path = Path(work_dir)
    work_dir_path.mkdir(parents=True, exist_ok=True)

    prelim_out = str(work_dir_path / f"{Path(matrix_file).stem}_estimate_prelim.nii.gz")

    flirt = fsl.FLIRT()
    flirt.inputs.in_file = moving_file
    flirt.inputs.reference = reference_file
    flirt.inputs.out_file = prelim_out
    flirt.inputs.out_matrix_file = matrix_file
    flirt.inputs.dof = int(dof)
    flirt.inputs.cost = cost
    flirt.inputs.interp = "trilinear"
    flirt.inputs.searchr_x = [-180, 180]
    flirt.inputs.searchr_y = [-180, 180]
    flirt.inputs.searchr_z = [-180, 180]

    estimate_command = flirt.cmdline
    run_tracked_command(estimate_command, work_dir, runtime)

    return {
        "estimate_command": estimate_command,
        "matrix": matrix_file_to_list(matrix_file),
        "estimated_output": prelim_out,
    }


def apply_existing_transform(
    moving_file: str,
    reference_file: str,
    output_file: str,
    matrix_file: str,
    work_dir: str,
    runtime: RuntimeControl,
) -> str:
    """
    Applicerar en redan beräknad FLIRT-transform på en bild.

    Detta steg skapar den slutliga coregistrerade NIfTI-filen i T1-space.
    """

    applyxfm = fsl.ApplyXFM()
    applyxfm.inputs.in_file = moving_file
    applyxfm.inputs.reference = reference_file
    applyxfm.inputs.in_matrix_file = matrix_file
    applyxfm.inputs.out_file = output_file
    applyxfm.inputs.apply_xfm = True
    applyxfm.inputs.interp = "trilinear"

    apply_command = applyxfm.cmdline
    run_tracked_command(apply_command, work_dir, runtime)
    return apply_command


def apply_transform_to_4d_by_volume(
    moving_4d_file: str,
    reference_file: str,
    output_file: str,
    matrix_file: str,
    work_dir: str,
    runtime: RuntimeControl,
) -> str:
    """
    Applicerar en FLIRT-transform på en 4D-bild volym för volym.

    Varje 3D-volym sparas temporärt, transformeras separat och slås sedan ihop
    igen med fslmerge. Om input redan är 3D används vanlig transformering.
    """

    if nib is None or np is None:
        raise RuntimeError("4D-transform kräver nibabel och numpy.")

    img = nib.load(moving_4d_file)
    data = np.asanyarray(img.dataobj)

    if data.ndim == 3:
        return apply_existing_transform(
            moving_file=moving_4d_file,
            reference_file=reference_file,
            output_file=output_file,
            matrix_file=matrix_file,
            work_dir=work_dir,
            runtime=runtime,
        )

    if data.ndim != 4:
        raise RuntimeError(f"Perfusionsfilen är varken 3D eller 4D: {moving_4d_file}")

    split_dir = Path(work_dir) / "apply_4d_split"
    split_dir.mkdir(parents=True, exist_ok=True)

    out_volumes: List[Path] = []
    apply_commands: List[str] = []

    for vol_idx in range(data.shape[3]):
        if runtime.cancel_event.is_set():
            raise UserCancelledError("Körningen avbröts av användaren.")

        in_vol = split_dir / f"vol_{vol_idx:04d}.nii.gz"
        out_vol = split_dir / f"vol_{vol_idx:04d}_xfm.nii.gz"

        vol_data = data[..., vol_idx]
        vol_img = nib.Nifti1Image(vol_data, img.affine, img.header.copy())
        nib.save(vol_img, str(in_vol))

        apply_command = apply_existing_transform(
            moving_file=str(in_vol),
            reference_file=reference_file,
            output_file=str(out_vol),
            matrix_file=matrix_file,
            work_dir=work_dir,
            runtime=runtime,
        )

        apply_commands.append(apply_command)
        out_volumes.append(out_vol)

    merge_command = "fslmerge -t " + " ".join(
        [shlex.quote(output_file)] + [shlex.quote(str(p)) for p in out_volumes]
    )
    run_tracked_command(merge_command, work_dir, runtime)

    return "\n".join(apply_commands + ["", "MergeCommand:", merge_command])


def run_flirt_coregistration(
    moving_file: str,
    reference_file: str,
    output_file: str,
    matrix_file: str,
    log_file: str,
    dof: int,
    cost: str,
    work_dir: str,
    runtime: RuntimeControl,
) -> Dict[str, Any]:
    """
    Applicerar en FLIRT-transform på en 4D-bild volym för volym.

    Varje 3D-volym sparas temporärt, transformeras separat och slås sedan ihop
    igen med fslmerge. Om input redan är 3D används vanlig transformering.
    """

    estimate_meta = estimate_flirt_transform(
        moving_file=moving_file,
        reference_file=reference_file,
        matrix_file=matrix_file,
        dof=dof,
        cost=cost,
        work_dir=work_dir,
        runtime=runtime,
    )

    apply_command = apply_existing_transform(
        moving_file=moving_file,
        reference_file=reference_file,
        output_file=output_file,
        matrix_file=matrix_file,
        work_dir=work_dir,
        runtime=runtime,
    )

    write_command_log(
        log_file=log_file,
        moving_file=moving_file,
        reference_file=reference_file,
        output_file=output_file,
        matrix_file=matrix_file,
        cost=cost,
        dof=dof,
        estimate_command=estimate_meta["estimate_command"],
        apply_command=apply_command,
    )

    return {
        "estimate_command": estimate_meta["estimate_command"],
        "apply_command": apply_command,
        "matrix": estimate_meta["matrix"],
        "estimated_output": estimate_meta["estimated_output"],
        "log_file": log_file,
    }


def process_one_registration(
    root_dir: Path,
    session_dir: Path,
    subject: str,
    session: str,
    moving_path: Path,
    reference_path: Path,
    moving_role: str,
    reference_role: str,
    cost: str,
    dof: int,
    work_dir: Path,
    runtime: RuntimeControl,
    output_subdir: Optional[Path] = None,
    output_name_suffix: str = "",
) -> Dict[str, Any]:
    """
    Kör en komplett coregistrering för en bild mot T1.

    Funktionen:
    1. Skapar output-sökvägar.
    2. Hoppar över registreringen om färdiga filer redan finns.
    3. Kör FLIRT för att estimera och applicera transformationen.
    4. Läser input- och outputmetadata.
    5. Beräknar similarity metrics.
    6. Sparar en JSON-sidecar med metadata och registreringsinformation.
    """

    out_nifti, out_json, out_mat, out_log = path_for_generated_output(
        root_dir=root_dir,
        session_dir=session_dir,
        subject=subject,
        session=session,
        moving_path=moving_path,
        output_subdir=output_subdir,
        name_suffix=output_name_suffix,
    )
    if runtime.cancel_event.is_set():
        raise UserCancelledError("Körningen avbröts av användaren.")

    existing_complete = (
        out_nifti.exists()
        and out_json.exists()
        and out_mat.exists()
        and out_log.exists()
    )

    if existing_complete and not OVERWRITE_EXISTING:

        similarity_metrics = compute_similarity_metrics_safe(
            reference_file=str(reference_path),
            registered_file=str(out_nifti),
        )

        return {
            "status": "skipped_existing",
            "moving_file": str(moving_path),
            "reference_file": str(reference_path),
            "output_file": str(out_nifti),
            "output_json": str(out_json),
            "transform_file": str(out_mat),
            "command_log_file": str(out_log),
            "cost_function": cost,
            "dof": int(dof),
            "similarity_metrics": similarity_metrics,
        }

    source_json_path = companion_json_path(moving_path)
    source_json = load_json_lenient(str(source_json_path)) if source_json_path else {}

    flirt_meta = run_flirt_coregistration(
        moving_file=str(moving_path),
        reference_file=str(reference_path),
        output_file=str(out_nifti),
        matrix_file=str(out_mat),
        log_file=str(out_log),
        dof=dof,
        cost=cost,
        work_dir=str(work_dir),
        runtime=runtime,
    )

    input_info = nifti_info(str(moving_path))
    output_info = nifti_info(str(out_nifti))

    similarity_metrics = compute_similarity_metrics_safe(
        reference_file=str(reference_path),
        registered_file=str(out_nifti),
    )

    reg_block = registration_json_block(
        moving_role=moving_role,
        reference_role=reference_role,
        matrix_path=str(out_mat),
        output_path=str(out_nifti),
        dof=dof,
        cost=cost,
        flirt_meta=flirt_meta,
    )
    reg_block["SimilarityMetrics"] = similarity_metrics

    output_json = build_output_json(
        source_json=source_json,
        input_info=input_info,
        output_info=output_info,
        modality_kind=moving_role,
        input_nifti_path=str(moving_path),
        output_nifti_path=str(out_nifti),
        reference_nifti_path=str(reference_path),
        input_json_path=str(source_json_path) if source_json_path else None,
        output_json_path=str(out_json),
        registration_meta=reg_block,
    )
    save_json(str(out_json), output_json)

    return {
        "status": "processed",
        "moving_file": str(moving_path),
        "reference_file": str(reference_path),
        "output_file": str(out_nifti),
        "output_json": str(out_json),
        "transform_file": str(out_mat),
        "command_log_file": str(out_log),
        "cost_function": cost,
        "dof": int(dof),
        "similarity_metrics": similarity_metrics,
    }


def process_perfusion_registration(
    root_dir: Path,
    session_dir: Path,
    subject: str,
    session: str,
    perfusion_path: Path,
    reference_path: Path,
    cost: str,
    dof: int,
    work_dir: Path,
    runtime: RuntimeControl,
    ref_method: str = "mean",
) -> Dict[str, Any]:
    """
    Kör coregistrering av perfusionsdata mot T1.

    Om perfusionsfilen är 4D skapas först en 3D-referensvolym, oftast som
    medelvärdet över tid. Denna referens registreras sedan till T1.
    """

    out_nifti, out_json, out_mat, out_log = path_for_generated_output(
        root_dir=root_dir,
        session_dir=session_dir,
        subject=subject,
        session=session,
        moving_path=perfusion_path,
        output_subdir=Path("perf"),
        name_suffix=f"cost-{cost}",
    )

    if runtime.cancel_event.is_set():
        raise UserCancelledError("Körningen avbröts av användaren.")

    existing_complete = (
        out_nifti.exists()
        and out_json.exists()
        and out_mat.exists()
        and out_log.exists()
    )

    if existing_complete and not OVERWRITE_EXISTING:
        similarity_metrics = compute_similarity_metrics_safe(
            reference_file=str(reference_path),
            registered_file=str(out_nifti),
        )

        return {
            "status": "skipped_existing",
            "moving_file": str(perfusion_path),
            "reference_file": str(reference_path),
            "output_file": str(out_nifti),
            "output_json": str(out_json),
            "transform_file": str(out_mat),
            "command_log_file": str(out_log),
            "cost_function": cost,
            "dof": int(dof),
            "similarity_metrics": similarity_metrics,
        }

    perfusion_ref_path = extract_perfusion_reference_volume(
        perf_path=perfusion_path,
        out_dir=work_dir / "ref_extract",
        method=ref_method,
    )

    source_json_path = companion_json_path(perfusion_path)
    source_json = load_json_lenient(str(source_json_path)) if source_json_path else {}

    estimate_meta = estimate_flirt_transform(
        moving_file=str(perfusion_ref_path),
        reference_file=str(reference_path),
        matrix_file=str(out_mat),
        dof=dof,
        cost=cost,
        work_dir=str(work_dir),
        runtime=runtime,
    )

    apply_command = apply_existing_transform(
        moving_file=str(perfusion_ref_path),
        reference_file=str(reference_path),
        output_file=str(out_nifti),
        matrix_file=str(out_mat),
        work_dir=str(work_dir),
        runtime=runtime,
    )

    write_command_log(
        log_file=str(out_log),
        moving_file=str(perfusion_path),
        reference_file=str(reference_path),
        output_file=str(out_nifti),
        matrix_file=str(out_mat),
        cost=cost,
        dof=dof,
        estimate_command=estimate_meta["estimate_command"],
        apply_command=apply_command,
    )

    input_info = nifti_info(str(perfusion_ref_path))
    output_info = nifti_info(str(out_nifti))

    similarity_metrics = compute_similarity_metrics_safe(
        reference_file=str(reference_path),
        registered_file=str(out_nifti),
    )

    flirt_meta = {
        "estimate_command": estimate_meta["estimate_command"],
        "apply_command": apply_command,
        "matrix": estimate_meta["matrix"],
    }

    reg_block = registration_json_block(
        moving_role=PERFUSION_LABEL,
        reference_role="MRI_T1",
        matrix_path=str(out_mat),
        output_path=str(out_nifti),
        dof=dof,
        cost=cost,
        flirt_meta=flirt_meta,
    )
    reg_block["SimilarityMetrics"] = similarity_metrics

    output_json = build_output_json(
        source_json=source_json,
        input_info=input_info,
        output_info=output_info,
        modality_kind=PERFUSION_LABEL,
        input_nifti_path=str(perfusion_path),
        output_nifti_path=str(out_nifti),
        reference_nifti_path=str(reference_path),
        input_json_path=str(source_json_path) if source_json_path else None,
        output_json_path=str(out_json),
        registration_meta=reg_block,
    )
    output_json["PerfusionReferenceVolumeFile"] = str(perfusion_ref_path)
    output_json["PerfusionReferenceVolumeMethod"] = ref_method

    save_json(str(out_json), output_json)

    return {
        "status": "processed",
        "moving_file": str(perfusion_path),
        "reference_file": str(reference_path),
        "output_file": str(out_nifti),
        "output_json": str(out_json),
        "transform_file": str(out_mat),
        "command_log_file": str(out_log),
        "cost_function": cost,
        "dof": int(dof),
        "similarity_metrics": similarity_metrics,
    }


def process_session(
    root_dir: Path,
    session_dir: Path,
    default_cost: str,
    diffusion_cost: str,
    perfusion_cost: str,
    dof: int,
    work_root: Path,
    run_t2: bool,
    run_diffusion: bool,
    run_perfusion: bool,
    use_contrast: bool,
    runtime: RuntimeControl,
) -> Dict[str, Any]:
    """
    Bearbetar en hel subject/session.

    PET till T1 körs alltid. T2, diffusion och perfusion körs bara om användaren
    har markerat respektive alternativ i GUI:t. Alla resultat samlas i en dict
    som senare används för batchsammanfattningen och Excel-exporten.
    """

    session_started_at = datetime.now()
    session_start_perf = time.perf_counter()

    subject = session_dir.parent.name
    session = session_dir.name

    if runtime.cancel_event.is_set():
        raise UserCancelledError("Körningen avbröts av användaren.")

    session_work = work_root / subject / session
    session_work.mkdir(parents=True, exist_ok=True)

    t1_path = find_t1_file(session_dir, use_contrast=use_contrast)
    pet_path = find_pet_file(session_dir)
    t2_path = find_t2_file(session_dir, use_contrast=use_contrast)
    diffusion_path = find_diffusion_file(session_dir)
    perfusion_path = find_perfusion_file(session_dir)

    if t1_path is None:
        raise FileNotFoundError(f"Hittade ingen T1-fil i {session_dir}")
    if pet_path is None:
        raise FileNotFoundError(f"Hittade ingen PET-fil i {session_dir}")

    results: Dict[str, Any] = {
        "subject": subject,
        "session": session,
        "session_dir": str(session_dir),
        "t1_reference": str(t1_path),
        "t2_source": str(t2_path) if t2_path else None,
        "diffusion_source": str(diffusion_path) if diffusion_path else None,
        "perfusion_source": str(perfusion_path) if perfusion_path else None,
        "run_t2": bool(run_t2),
        "run_diffusion": bool(run_diffusion),
        "run_perfusion": bool(run_perfusion),
        "use_contrast": bool(use_contrast),
        "default_cost": default_cost,
        "diffusion_cost": diffusion_cost,
        "perfusion_cost": perfusion_cost,
        "pet": None,
        "t2": None,
        "diffusion": None,
        "perfusion": None,
    }

    results["pet"] = process_one_registration(
        root_dir=root_dir,
        session_dir=session_dir,
        subject=subject,
        session=session,
        moving_path=pet_path,
        reference_path=t1_path,
        moving_role="PET",
        reference_role="MRI_T1",
        cost=default_cost,
        dof=dof,
        work_dir=session_work / "pet_to_t1",
        runtime=runtime,
        output_name_suffix=f"cost-{default_cost}",
    )

    if run_t2 and t2_path is not None:
        results["t2"] = process_one_registration(
            root_dir=root_dir,
            session_dir=session_dir,
            subject=subject,
            session=session,
            moving_path=t2_path,
            reference_path=t1_path,
            moving_role="MRI_T2",
            reference_role="MRI_T1",
            cost=default_cost,
            dof=dof,
            work_dir=session_work / "t2_to_t1",
            runtime=runtime,
            output_name_suffix=f"cost-{default_cost}",
        )

    if run_diffusion and diffusion_path is not None:
        diffusion_b0_path = extract_b0_volume(
            dwi_path=diffusion_path,
            out_dir=session_work / "diffusion_to_t1" / "b0_extract",
        )

        results["diffusion"] = process_one_registration(
            root_dir=root_dir,
            session_dir=session_dir,
            subject=subject,
            session=session,
            moving_path=diffusion_b0_path,
            reference_path=t1_path,
            moving_role="MRI_DWI_B0",
            reference_role="MRI_T1",
            cost=diffusion_cost,
            dof=dof,
            work_dir=session_work / "diffusion_to_t1",
            runtime=runtime,
            output_subdir=Path("dwi"),
            output_name_suffix=f"cost-{diffusion_cost}",
        )

    if run_perfusion and perfusion_path is not None:
        results["perfusion"] = process_perfusion_registration(
            root_dir=root_dir,
            session_dir=session_dir,
            subject=subject,
            session=session,
            perfusion_path=perfusion_path,
            reference_path=t1_path,
            cost=perfusion_cost,
            dof=dof,
            work_dir=session_work / "perfusion_to_t1",
            runtime=runtime,
            ref_method="mean",
        )

    session_finished_at = datetime.now()
    session_elapsed = time.perf_counter() - session_start_perf

    results["started_at"] = session_started_at.isoformat(timespec="seconds")
    results["finished_at"] = session_finished_at.isoformat(timespec="seconds")
    results["elapsed_seconds"] = round(session_elapsed, 3)
    results["elapsed_time"] = format_duration(session_elapsed)

    return results


def collect_t1_space_files_for_fsl(root_dir: Path, session_dir: Path) -> List[Path]:
    """
    Samlar de filer som ska öppnas tillsammans i FSL.

    Listan innehåller T1-referensen och alla coregistrerade filer i T1-space.
    """

    files: List[Path] = []

    t1_path = find_reference_t1_for_session(root_dir, session_dir)
    if t1_path is not None:
        files.append(t1_path)

    files.extend(find_coreg_files_for_session(root_dir, session_dir))

    return sorted({p.resolve() for p in files}, key=session_view_priority)


def is_checkerboard_qc_file(path: Path) -> bool:
    """
    Kontrollerar om en fil är en checkerboard-QC-fil.
    """

    return CHECKERBOARD_MARKER in path.name.lower()


def _to_3d_checkerboard_data(path: str) -> Tuple[np.ndarray, str]:
    """
    Returnerar 3D-data för checkerboard-QC utan att ladda hela 4D-bilder i RAM.

    3D-bild: läses som float32.
    4D-bild: endast första volymen används, alltså data[..., 0].
    """

    img = nib.load(path)
    shape = img.shape

    if len(shape) == 3:
        data = np.asarray(img.dataobj, dtype=np.float32)
        return data, "3D"

    if len(shape) == 4:
        if shape[3] < 1:
            raise RuntimeError(f"4D-bilden har inga volymer: {path}")

        # Läser bara första volymen, inte hela 4D-filen.
        data = np.asarray(img.dataobj[..., 0], dtype=np.float32)
        return data, "4D_first_volume"

    raise RuntimeError(f"Kan inte skapa checkerboard för {len(shape)}D-bild: {path}")


def _robust_normalize_for_qc(data: np.ndarray) -> np.ndarray:
    """
    Normaliserar intensiteter robust för visuell jämförelse.

    Percentilerna 2 och 98 används för att minska påverkan från extrema värden.
    Detta gör att bilder med olika intensitetsskalor, till exempel PET och T1,
    kan visas tillsammans på ett mer jämförbart sätt.
    """

    out = np.zeros(data.shape, dtype=np.float32)
    finite = np.isfinite(data)

    if not np.any(finite):
        return out

    mask = finite & (data != 0)
    values = data[mask]

    if values.size < 10:
        values = data[finite]

    if values.size == 0:
        return out

    low, high = np.percentile(values, [2, 98])

    if not np.isfinite(low) or not np.isfinite(high) or high <= low:
        low = float(np.min(values))
        high = float(np.max(values))

    if high <= low:
        return out

    out[finite] = np.clip((data[finite] - low) / (high - low), 0, 1) * 1000.0
    return out


def create_checkerboard_image(
    reference_file: str,
    registered_file: str,
    output_file: str,
    block_size: int = CHECKERBOARD_BLOCK_SIZE,
) -> Dict[str, Any]:
    """
    Skapar en checkerboard-bild för visuell kontroll av registreringen.

    Bilden växlar mellan block från T1-referensen och den coregistrerade bilden.
    Om registreringen är bra ska anatomiska strukturer ligga rätt över
    blockgränserna.
    
    reference_file ska vara T1.
    registered_file ska redan ligga i T1-space.
    """

    if nib is None or np is None:
        raise RuntimeError("Checkerboard QC kräver nibabel och numpy.")

    block_size = max(1, int(block_size))

    reference_img = nib.load(reference_file)

    reference_data, reference_kind = _to_3d_checkerboard_data(reference_file)
    registered_data, registered_kind = _to_3d_checkerboard_data(registered_file)

    if reference_data.shape != registered_data.shape:
        raise RuntimeError(
            "T1 och coregistrerad bild har olika shape. "
            f"T1: {reference_data.shape}, coreg: {registered_data.shape}. "
            "Checkerboard kräver att bilderna redan ligger på samma grid."
        )

    reference_show = _robust_normalize_for_qc(reference_data)
    registered_show = _robust_normalize_for_qc(registered_data)

    reference_negated = False
    if REFERENCE_NEGATED:
        # Gör T1  negativ i checkerboard-filen.
        # Bakgrund hålls kvar som 0.
        ref_mask = np.isfinite(reference_show) 

        negative_reference_show = np.zeros_like(reference_show, dtype=np.float32)
        negative_reference_show[ref_mask] = -1.0 * reference_show[ref_mask]

        reference_show = negative_reference_show
        reference_negated = True

    nx, ny, nz = reference_show.shape

    x_blocks = np.arange(nx)[:, None, None] // block_size
    y_blocks = np.arange(ny)[None, :, None] // block_size
    z_blocks = np.arange(nz)[None, None, :] // block_size

    checker_mask = ((x_blocks + y_blocks + z_blocks) % 2) == 0

    checker_data = np.where(
        checker_mask,
        reference_show,
        registered_show,
    ).astype(np.float32)

    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    out_header = reference_img.header.copy()
    out_header.set_data_dtype(np.float32)

    out_img = nib.Nifti1Image(
        checker_data,
        reference_img.affine,
        out_header,
    )

    nib.save(out_img, str(output_path))

    return {
        "computed": True,
        "reference_file": str(reference_file),
        "registered_file": str(registered_file),
        "checkerboard_file": str(output_file),
        "block_size_voxels": int(block_size),
        "reference_data_kind": reference_kind,
        "registered_data_kind": registered_kind,
        "reference_negated": reference_negated,
        "shape": [int(v) for v in checker_data.shape],
        "intensity_scaling": "robust_2_98_percentile_per_image_to_0_1000",
    }


def checkerboard_output_paths_for_coreg(
    root_dir: Path,
    session_dir: Path,
    coreg_file: Path,
) -> Tuple[Path, Path]:
    """
    Skapar output-sökvägar för en checkerboard-QC-fil och dess JSON-sidecar.
    """

    subject = session_dir.parent.name
    session = session_dir.name

    qc_dir = (
        root_dir
        / "derivatives"
        / "registration"
        / CHECKERBOARD_QC_DIRNAME
        / subject
        / session
    )
    qc_dir.mkdir(parents=True, exist_ok=True)

    base = basename_without_known_suffixes(coreg_file)
    out_base = f"{base}{CHECKERBOARD_FILENAME_SUFFIX}"

    out_nifti = qc_dir / f"{out_base}.nii.gz"
    out_json = qc_dir / f"{out_base}.json"

    return out_nifti, out_json


def create_checkerboards_for_session(
    root_dir: Path,
    session_dir: Path,
    block_size: int = CHECKERBOARD_BLOCK_SIZE,
) -> List[Dict[str, Any]]:
    """
    Skapar checkerboard-QC för alla coregistrerade filer i en session.

    Varje coregistrerad bild jämförs visuellt mot T1-referensen.
    """

    reference_file = find_reference_t1_for_session(root_dir, session_dir)

    if reference_file is None:
        raise FileNotFoundError(f"Hittade ingen T1-reference för sessionen: {session_dir}")

    coreg_files = [
        p for p in find_coreg_files_for_session(root_dir, session_dir)
        if not is_checkerboard_qc_file(p)
    ]

    if not coreg_files:
        raise FileNotFoundError(
            f"Hittade inga färdiga coregistrerade filer för sessionen: {session_dir}"
        )

    results: List[Dict[str, Any]] = []

    for coreg_file in coreg_files:
        out_nifti, out_json = checkerboard_output_paths_for_coreg(
            root_dir=root_dir,
            session_dir=session_dir,
            coreg_file=coreg_file,
        )

        if out_nifti.exists() and out_json.exists():
            results.append({
                "computed": False,
                "status": "skipped_existing",
                "reference_file": str(reference_file),
                "registered_file": str(coreg_file),
                "checkerboard_file": str(out_nifti),
                "checkerboard_json": str(out_json),
                "block_size_voxels": int(block_size),
                "reference_negated": bool(REFERENCE_NEGATED),
            })
            continue

        meta = create_checkerboard_image(
            reference_file=str(reference_file),
            registered_file=str(coreg_file),
            output_file=str(out_nifti),
            block_size=block_size,
        )

        sidecar = {
            "ProgramName": APP_TITLE,
            "CreatedAt": datetime.now().isoformat(timespec="seconds"),
            "QCType": "checkerboard",
            "ReferenceFile": str(reference_file),
            "RegisteredFile": str(coreg_file),
            "CheckerboardFile": str(out_nifti),
            "CheckerboardJSONFile": str(out_json),
            "BlockSizeVoxels": int(block_size),
            "ReferenceNegated": meta.get("reference_negated"),
            "IntensityScaling": meta.get("intensity_scaling"),
            "ReferenceDataKind": meta.get("reference_data_kind"),
            "RegisteredDataKind": meta.get("registered_data_kind"),
            "Shape": meta.get("shape"),
        }

        save_json(str(out_json), sidecar)

        meta["checkerboard_json"] = str(out_json)
        results.append(meta)

    return results


def find_checkerboard_qc_files_for_session(
    root_dir: Path,
    session_dir: Path,
) -> List[Path]:
    """
    Letar upp befintliga checkerboard-QC-filer för en session.
    """

    subject = session_dir.parent.name
    session = session_dir.name

    qc_dir = (
        root_dir
        / "derivatives"
        / "registration"
        / CHECKERBOARD_QC_DIRNAME
        / subject
        / session
    )

    if not qc_dir.exists():
        return []

    return sorted(
        p for p in qc_dir.rglob("*")
        if p.is_file() and is_nifti_file(str(p)) and is_checkerboard_qc_file(p)
    )


def find_registered_subject_session_dirs(root_dir: Path) -> List[Tuple[str, str, Path]]:
    """
    Hittar alla subject/session-mappar där färdiga coregistreringsfiler finns.
    """

    found: List[Tuple[str, str, Path]] = []

    for subject, session, session_dir in find_subject_session_dirs(root_dir):
        if find_coreg_files_for_session(root_dir, session_dir):
            found.append((subject, session, session_dir))

    return found


def create_checkerboards_for_all_registered_sessions(
    root_dir: Path,
    block_size: int = CHECKERBOARD_BLOCK_SIZE,
) -> List[Dict[str, Any]]:
    """
    Skapar checkerboard-QC för alla sessioner som har färdiga coregistreringar.

    Fel i en session stoppar inte hela processen, utan sparas i resultatlistan.
    """

    sessions = find_registered_subject_session_dirs(root_dir)

    if not sessions:
        raise RuntimeError(
            "Hittade inga sessioner med färdiga coregistreringar."
        )

    all_results: List[Dict[str, Any]] = []

    for subject, session, session_dir in sessions:
        session_result: Dict[str, Any] = {
            "subject": subject,
            "session": session,
            "session_dir": str(session_dir),
            "status": None,
            "checkerboard_count": 0,
            "checkerboards": [],
            "error": None,
        }

        try:
            checkerboards = create_checkerboards_for_session(
                root_dir=root_dir,
                session_dir=session_dir,
                block_size=block_size,
            )

            session_result["status"] = "processed"
            session_result["checkerboard_count"] = len(checkerboards)
            session_result["checkerboards"] = checkerboards

        except Exception as exc:
            session_result["status"] = "failed"
            session_result["error"] = str(exc)
            session_result["traceback"] = traceback.format_exc()

        all_results.append(session_result)

    return all_results


class CoregBatchApp(tk.Tk):
    """
    Tkinter-baserat grafiskt gränssnitt för batch-coregistrering.

    Klassen bygger hela användargränssnittet, hanterar knapptryckningar,
    startar bakgrundstrådar och tar emot statusuppdateringar via en kö.
    Tunga beräkningar körs inte direkt i GUI-tråden, eftersom det skulle göra
    att fönstret fryser.
    """

    def __init__(self):
        """
        Initierar huvudfönstret, Tkinter-variabler, trådhantering och startvyn.

        Här skapas även kön som används för kommunikation mellan worker-trådar
        och GUI-tråden.
        """

        super().__init__()
        self.use_contrast_var = tk.BooleanVar(value=False)
        self.diffusion_cost_var = tk.StringVar(value=DEFAULT_DIFFUSION_COST)
        self.perfusion_cost_var = tk.StringVar(value=DEFAULT_PERFUSION_COST)

        self.title(APP_TITLE)
        self.geometry("1100x800")

        self.cancel_event = threading.Event()
        self.close_when_done = False
        self.runtime_control: Optional[RuntimeControl] = None
        self.executor: Optional[ThreadPoolExecutor] = None
        self.busy_progress_convert = None
        self.other_log_widget = None
        self.comob = None

        self.protocol("WM_DELETE_WINDOW", self.request_exit)

        self.available_cpu_cores = get_available_cpu_count()
        self.recommended_cpu_workers = get_recommended_worker_count()
        self.worker_options = get_worker_options(self.available_cpu_cores)

        self.max_workers_var = tk.StringVar(value=str(self.recommended_cpu_workers))
        self.converter_workers_var = tk.StringVar(value=str(self.recommended_cpu_workers))

        self.cpu_info_text = tk.StringVar(
            value=(
                f"Tillgängliga CPU-kärnor: {self.available_cpu_cores}. "
                f"Rekommenderat: {self.recommended_cpu_workers}."
            )
        )

        self.max_workers_combo = None

        self.root_dir = tk.StringVar()
        self.status_text = tk.StringVar(
            value="Välj rotmappen som innehåller sub-*/ses-* och tryck Run."
        )
        self.current_item_text = tk.StringVar(value="Ingen körning aktiv.")
        self.progress_value = tk.DoubleVar(value=0.0)

        self.run_t2_var = tk.BooleanVar(value=False)
        self.run_diffusion_var = tk.BooleanVar(value=False)
        self.run_perfusion_var = tk.BooleanVar(value=False)
        self.validate_bids_var = tk.BooleanVar(value=VALIDATE_BIDS_BEFORE_RUN_DEFAULT)

        self.session_choice = tk.StringVar(value="")
        self.session_map: Dict[str, str] = {}

        self.other_group_root_dir = tk.StringVar()
        self.other_group_output_dir = tk.StringVar()

        self.current_view = "coregistrering_ui"

        self.worker_thread: Optional[threading.Thread] = None
        self.ui_queue: "queue.Queue[Tuple[str, Any]]" = queue.Queue()
        self.is_running = False

        self.content_frame: Optional[ttk.Frame] = None

        self.log_widget = None
        self.session_progress = None
        self.busy_progress = None
        self.run_button = None
        self.clear_button = None
        self.run_t2_check = None
        self.run_diffusion_check = None
        self.run_perfusion_check = None
        self.use_contrast_check = None
        self.diffusion_cost_combo = None
        self.perfusion_cost_combo = None
        self.session_combo = None
        self.refresh_sessions_button = None
        self.open_session_button = None
        self.validate_bids_check = None

        ###___Converter Variables_______________
        self.Convert_Thread_Active = False
        self.Breaking_Converter_Program = False
        self.Run_in_name_bool = tk.BooleanVar(value=False)
        self.guarantee_dcm = tk.BooleanVar(value=True)
        self.YN = tk.BooleanVar(value=True)
        self.other_status_text = tk.StringVar(
            value="Välj rotmapp med DICOM-filer, välj map där BIDS-struktur önskas, och tryck Run."
        )
        ###________________________________

        self._build_menu()

        self.content_frame = ttk.Frame(self)
        self.content_frame.pack(fill="both", expand=True)

        if running_on_windows():
            self.show_BIDS_Converter()
            self.other_status_text.set(
                "Programmet körs i Windows. Coregistrering med FSL är avstängd, därför öppnas BIDS Converter automatiskt."
            )
        else:
            self.show_coregistrering_ui()

        self.after(150, self.process_ui_queue)
        
    def _build_menu(self) -> None:
        """
        Skapar programmets menyfält.

        Menyn innehåller val för olika vyer, similarity metrics, checkerboard-QC
        och avslutning av programmet.
        """

        menubar = tk.Menu(self)

        view_menu = tk.Menu(menubar, tearoff=0)
        view_menu.add_command(label="Coregistrering UI", command=self.show_coregistrering_ui)
        view_menu.add_command(label="BIDS Converter", command=self.show_BIDS_Converter)
        menubar.add_cascade(label="Mode", menu=view_menu)

        metrics_menu = tk.Menu(menubar, tearoff=0)
        metrics_menu.add_command(
            label="Skapa/uppdatera similarity metrics Excel",
            command=self.create_similarity_metrics_from_menu,
        )
        metrics_menu.add_command(
            label="Öppna similarity metrics Excel",
            command=self.open_similarity_metrics_from_menu,
        )
        metrics_menu.add_separator()
        metrics_menu.add_command(
            label="Skapa checkerboard QC för alla färdiga sessioner",
            command=self.create_checkerboard_all_from_menu,
        )
        metrics_menu.add_command(
            label="Skapa checkerboard QC för vald session",
            command=self.create_checkerboard_from_menu,
        )
        metrics_menu.add_command(
            label="Öppna checkerboard QC för vald session i FSL",
            command=self.open_checkerboard_from_menu,
        )
        menubar.add_cascade(label="Coregistrering QC", menu=metrics_menu)

        app_menu = tk.Menu(menubar, tearoff=0)
        app_menu.add_command(label="Välj/ändra FSL-sökväg", command=self.select_fsl_manually)
        app_menu.add_separator()
        app_menu.add_command(label="Exit", command=self.request_exit)
        menubar.add_cascade(label="App", menu=app_menu)

        self.config(menu=menubar)

    def _widget_exists(self, widget: Any) -> bool:
        """
        Kontrollerar om en Tkinter-widget finns och fortfarande är aktiv.
        """
        
        return widget is not None and bool(widget.winfo_exists())

    def _safe_configure(self, widget: Any, **kwargs: Any) -> None:
        """
        Ändrar inställningar på en widget endast om widgeten finns.

        Detta skyddar mot fel när GUI:t byter vy och gamla widgets har förstörts.
        """

        if self._widget_exists(widget):
            widget.configure(**kwargs)

    def _reset_view_widget_refs(self) -> None:
        """
        Nollställer referenser till widgets i den aktuella vyn.
        """
        
        self.log_widget = None
        self.session_progress = None
        self.busy_progress = None
        self.run_button = None
        self.clear_button = None
        self.run_t2_check = None
        self.run_diffusion_check = None
        self.run_perfusion_check = None
        self.use_contrast_check = None
        self.diffusion_cost_combo = None
        self.perfusion_cost_combo = None
        self.max_workers_combo = None
        self.session_combo = None
        self.refresh_sessions_button = None
        self.open_session_button = None
        self.validate_bids_check = None

    def _clear_current_ui(self) -> None:
        """
        Tar bort alla widgets från den nuvarande vyn.

        Används när programmet byter mellan huvudvyn och den BIDS Converter-vyn.
        """
        
        if self.content_frame is not None:
            for child in self.content_frame.winfo_children():
                child.destroy()
        self._reset_view_widget_refs()

    def request_exit(self) -> None:
        """
        Hanterar när användaren vill stänga programmet.

        Om en körning pågår får användaren välja mellan att låta jobben bli klara,
        avbryta direkt eller fortsätta köra.
        """

        if not self.is_running:
            self.destroy()
            return

        answer = messagebox.askyesnocancel(
            "Stäng programmet",
            "Programmet kör fortfarande.\n\n"
            "Ja = låt pågående jobb bli klara och stäng programmet när allt är färdigt.\n"
            "Nej = avbryt alla pågående jobb direkt och stäng programmet nu.\n"
            "Avbryt = fortsätt köra programmet."
        )

        if answer is None:
            return

        if answer is True:
            self.close_when_done = True
            self.status_text.set("Programmet kommer att stängas automatiskt när alla pågående jobb är klara.")
            self.log("Användaren valde: slutför pågående jobb och stäng när klart.")
        else:
            self.log("Användaren valde: avbryt alla jobb och stäng nu.")
            self.abort_all_and_close()

    def abort_all_and_close(self) -> None:
        """
        Avbryter alla pågående jobb och stänger programmet.
        """

        self.cancel_event.set()

        if self.executor is not None:
            try:
                self.executor.shutdown(wait=False, cancel_futures=True)
            except Exception:
                pass

        if self.runtime_control is not None:
            try:
                self.runtime_control.terminate_all()
            except Exception:
                pass

        self.destroy()

    def show_coregistrering_ui(self) -> None:
        """
        Visar huvudgränssnittet för coregistrering.
        """

        if self.is_running:
            messagebox.showwarning("Körning pågår", "Byt inte UI medan programmet kör.")
            return

        self.current_view = "coregistrering_ui"
        self._clear_current_ui()
        self._build_coregistrering_ui()
        self.update_session_controls_state()

        if self.root_dir.get().strip():
            try:
                self.refresh_session_choices()
            except Exception:
                pass

    def show_BIDS_Converter(self) -> None:
        """
        Visar den alternativa placeholder-vyn för den andra gruppen.
        """
       
        if self.is_running:
            messagebox.showwarning("Körning pågår", "Byt inte UI medan programmet kör.")
            return

        self.current_view = "BIDS_convers"
        self._clear_current_ui()
        self._build_BIDS_Converter()

    def select_other_group_root_dir(self) -> None:
        """
        Öppnar en dialog för att välja rotmapp i placeholder-vyn.
        """

        if self.is_running:
            return

        path = filedialog.askdirectory(
            title="Välj rotmapp för BIDS Converter",
            initialdir=self.other_group_root_dir.get().strip() or "/",
        )
        if path:
            self.other_group_root_dir.set(path)

    def select_other_group_output_dir(self) -> None:
        """
        Öppnar en dialog för att välja rotmapp i placeholder-vyn.
        """
        if self.is_running:
            return

        path = filedialog.askdirectory(
            title="Välj outputmapp för BIDS Conveter",
            initialdir=self.other_group_output_dir.get().strip() or "/",
        )
        if path:
            self.other_group_output_dir.set(path)
            if self.current_view == "BIDS_convers":
                self.refresh_session_choices()

    def break_bids_converter(self):
        """Changes the break variable for converter program\n
        Group 13"""
        self.Breaking_Converter_Program = True
        if self.Convert_Thread_Active == True:
            self.update_other_status_text("Stoppar processen")

    def get_broken_state(self):
        """Return the bool for if the break button has been pressed\n
        The converter program checks this variable between loops and 
        breaks internaly if True. This allows quick killing of the program,
        while still doing the exiting procedures such as waiting for subprocess
        to close, saving data, and creating tables for what has been done\n
        Group 13"""
        return self.Breaking_Converter_Program

    def run_bids_converter_starter(self):
        """Starts a thread that starts the converter program thread\n
        I did it this way because i wanted to use .join to set active
        state to false\n
        Could be changed, it's not my darling, but as far as i know it 
        works without problems\n\n
        Group 13"""
        if self.Convert_Thread_Active == False:
            startThread = threading.Thread(target=self.run_bids_converter)
            startThread.daemon = True
            startThread.start()

    def run_bids_converter(self) -> None:
        """Starts a new thread with the converter program 
        if one is not currently running. Then waits for the 
        thread to join and opens it upp for activation\n 
        Group 13"""
        if self.Convert_Thread_Active == False:
            self.Convert_Thread_Active = True
            self.Breaking_Converter_Program = False
            self.set_busy_convert(self.Convert_Thread_Active)
            self.Convert_Thread = threading.Thread(
                    target=Conversion_Handler.prepAndConvert,
                    args=(
                        Path(self.other_group_root_dir.get().strip()),
                        Path(self.other_group_output_dir.get().strip()),
                        int(self.comob.get()),
                        self.YN.get(),
                        self.Run_in_name_bool.get(),
                        self.guarantee_dcm.get(),
                        self.update_other_status_text,
                        self.update_other_text_box,
                        self.get_broken_state))
            self.Convert_Thread.daemon = True
            self.Convert_Thread.start()
            self.Convert_Thread.join()
            self.Convert_Thread_Active = False
            self.set_busy_convert(self.Convert_Thread_Active)
            self.Breaking_Converter_Program = False

    def set_busy_convert(self, busy: bool) -> None:
        """Starts and ends the progressbar animation\n
        Group 13"""
        if self._widget_exists(self.busy_progress_convert):
            if busy == True:
                self.busy_progress_convert.start()

            elif busy == False:
                self.busy_progress_convert["value"] = 0

                self.busy_progress_convert.stop()

        self.update_idletasks()

    def update_other_status_text(self,thenewline:str):
        """Changes the text on the 'ny_ui' status label\n
        Group 13"""
        self.other_status_text.set(thenewline)

    def update_other_text_box(self, text:str):
        """Adds text to the 'ny_ui' text box\n
        Group 13"""
        if not self._widget_exists(self.other_log_widget):
            return
        self.other_log_widget.configure(state="normal")
        self.other_log_widget.insert("end", text + "\n")
        self.other_log_widget.see("end")
        self.other_log_widget.configure(state="disabled")
        self.update_idletasks()

    def _build_coregistrering_ui(self) -> None:
        """
        Bygger hela huvudgränssnittet för coregistrering.

        Här skapas fält för rotmapp, valfria registreringar, progressbar,
        sessionval, loggruta och kontrollknappar.
        """

        main = ttk.Frame(self.content_frame, padding=16)
        main.pack(fill="both", expand=True)
        main.columnconfigure(1, weight=1)
        main.rowconfigure(7, weight=1)

        ###___Styles________________________________
        ### Available on all platform: alt, clam, classic, default
        # Create a style
        style = ttk.Style(main)
        # Set the theme with the theme_use method
        style.theme_use('default')  # put the theme name here, that you want to use
        ###_________________________________________

        ttk.Label(
            main,
            text=APP_TITLE,
            font=("Segoe UI", 14, "bold"),
        ).grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 10))

        ttk.Label(
            main,
            text=(
                "Programmet frågar efter en rotmapp. Det söker igenom alla sub-*/ses-* och registrerar "
                "PET till T1. T2, diffusion och perfusion till T1 körs bara om du markerar respektive kryssruta. "
                "I FSL öppnas bara T1-bilden och filer som redan ligger i T1-space. "
                "Similarity metrics kan skapas via menyn längst upp."
            ),
            wraplength=1040,
        ).grid(row=1, column=0, columnspan=3, sticky="w", pady=(0, 14))

        ttk.Label(main, text="Rotmapp:").grid(row=2, column=0, sticky="w", pady=6)
        ttk.Entry(main, textvariable=self.root_dir).grid(row=2, column=1, sticky="ew", padx=8)
        ttk.Button(main, text="Browse", command=self.select_root_dir).grid(row=2, column=2, sticky="ew")

        options_frame = ttk.LabelFrame(main, text="Valfria coregistreringar", padding=10)
        options_frame.grid(row=3, column=0, columnspan=3, sticky="ew", pady=(14, 10))
        options_frame.columnconfigure(0, weight=1)
        options_frame.columnconfigure(1, weight=1)
        options_frame.columnconfigure(2, weight=1)
        options_frame.columnconfigure(3, weight=1)

        ttk.Label(options_frame, text="Parallella workers:").grid(
            row=1, column=0, sticky="w", pady=(10, 0)
        )

        self.run_t2_check = ttk.Checkbutton(
            options_frame,
            text="Kör T2 -> T1 coregistrering",
            variable=self.run_t2_var,
        )
        self.run_t2_check.grid(row=0, column=0, sticky="w", padx=(0, 12))

        self.run_diffusion_check = ttk.Checkbutton(
            options_frame,
            text="Kör diffusion -> T1 coregistrering",
            variable=self.run_diffusion_var,
        )
        self.run_diffusion_check.grid(row=0, column=1, sticky="w", padx=(0, 12))

        self.run_perfusion_check = ttk.Checkbutton(
            options_frame,
            text="Kör perfusion -> T1 coregistrering",
            variable=self.run_perfusion_var,
        )
        self.run_perfusion_check.grid(row=0, column=2, sticky="w", padx=(0, 12))

        self.use_contrast_check = ttk.Checkbutton(
            options_frame,
            text="Använd kontrast (ce-gd) för T1/T2",
            variable=self.use_contrast_var,
        )
        self.use_contrast_check.grid(row=0, column=3, sticky="w")

        self.validate_bids_check = ttk.Checkbutton(
            options_frame,
            text="Validera BIDS innan körning",
            variable=self.validate_bids_var,
        )
        
        self.validate_bids_check.grid(
            row=3,
            column=0,
            columnspan=4,
            sticky="w",
            pady=(12, 0),
        )

        self.max_workers_combo = ttk.Combobox(
            options_frame,
            textvariable=self.max_workers_var,
            values=self.worker_options,
            state="readonly",
            width=8,
        )
        self.max_workers_combo.grid(row=1, column=1, sticky="w", pady=(10, 0))

        ttk.Label(
            options_frame,
            textvariable=self.cpu_info_text,
        ).grid(row=1, column=2, columnspan=2, sticky="w", pady=(10, 0))

        ttk.Label(options_frame, text="Diffusion cost:").grid(
            row=2, column=0, sticky="w", pady=(10, 0)
        )

        self.diffusion_cost_combo = ttk.Combobox(
            options_frame,
            textvariable=self.diffusion_cost_var,
            values=COST_FUNCTION_OPTIONS,
            state="readonly",
            width=12,
        )
        self.diffusion_cost_combo.grid(row=2, column=1, sticky="w", pady=(10, 0))

        ttk.Label(options_frame, text="Perfusion cost:").grid(
            row=2, column=2, sticky="w", pady=(10, 0)
        )

        self.perfusion_cost_combo = ttk.Combobox(
            options_frame,
            textvariable=self.perfusion_cost_var,
            values=COST_FUNCTION_OPTIONS,
            state="readonly",
            width=12,
        )
        self.perfusion_cost_combo.grid(row=2, column=3, sticky="w", pady=(10, 0))

        btn_frame = ttk.Frame(main)
        btn_frame.grid(row=4, column=0, columnspan=3, sticky="w", pady=(0, 10))

        self.run_button = ttk.Button(btn_frame, text="Run", command=self.run_processing)
        self.run_button.pack(side="left", padx=(0, 8))

        self.clear_button = ttk.Button(btn_frame, text="Clear", command=self.clear_fields)
        self.clear_button.pack(side="left", padx=(0, 8))

        ttk.Button(btn_frame, text="Exit", command=self.request_exit).pack(side="left")

        session_box = ttk.LabelFrame(main, text="Öppna session i FSL", padding=10)
        session_box.grid(row=5, column=0, columnspan=3, sticky="ew", pady=(0, 10))
        session_box.columnconfigure(0, weight=1)

        ttk.Label(
            session_box,
            text=(
                "Efter körning visas bara subject/session där det finns färdig coregistrering. "
                "När du öppnar i FSL laddas bara T1 + coreg-filer."
            ),
            wraplength=1040,
        ).grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 8))

        self.session_combo = ttk.Combobox(
            session_box,
            textvariable=self.session_choice,
            state="disabled",
        )
        self.session_combo.grid(row=1, column=0, sticky="ew", padx=(0, 8))

        self.refresh_sessions_button = ttk.Button(
            session_box,
            text="Läs in färdiga sessioner",
            command=self.refresh_session_choices,
        )
        self.refresh_sessions_button.grid(row=1, column=1, sticky="ew", padx=(0, 8))

        self.open_session_button = ttk.Button(
            session_box,
            text="Öppna T1 + coreg i FSL",
            command=self.open_selected_session_in_fsl,
            state="disabled",
        )
        self.open_session_button.grid(row=1, column=2, sticky="ew")

        progress_box = ttk.LabelFrame(main, text="Arbetsstatus", padding=10)
        progress_box.grid(row=6, column=0, columnspan=3, sticky="ew", pady=(0, 10))
        progress_box.columnconfigure(0, weight=1)

        ttk.Label(progress_box, textvariable=self.current_item_text, wraplength=1040).grid(
            row=0, column=0, sticky="w", pady=(0, 8)
        )

        self.session_progress = ttk.Progressbar(
            progress_box,
            orient="horizontal",
            mode="determinate",
            maximum=100,
            variable=self.progress_value,
        )
        self.session_progress.grid(row=1, column=0, sticky="ew", pady=(0, 8))

        ttk.Label(progress_box, text="Pågående registrering:").grid(
            row=2, column=0, sticky="w", pady=(0, 4)
        )

        self.busy_progress = ttk.Progressbar(
            progress_box,
            orient="horizontal",
            mode="indeterminate",
        )
        self.busy_progress.grid(row=3, column=0, sticky="ew")

        status_box = ttk.LabelFrame(main, text="Status", padding=10)
        status_box.grid(row=7, column=0, columnspan=3, sticky="nsew")
        status_box.columnconfigure(0, weight=1)
        status_box.rowconfigure(1, weight=1)

        ttk.Label(status_box, textvariable=self.status_text, wraplength=1040).grid(
            row=0, column=0, sticky="w", pady=(0, 10)
        )

        log_frame = ttk.Frame(status_box)
        log_frame.grid(row=1, column=0, sticky="nsew")
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)

        self.log_widget = tk.Text(log_frame, wrap="word", height=24)
        self.log_widget.grid(row=0, column=0, sticky="nsew")

        log_scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_widget.yview)
        log_scrollbar.grid(row=0, column=1, sticky="ns")

        self.log_widget.configure(yscrollcommand=log_scrollbar.set, state="disabled")

    def _build_BIDS_Converter(self) -> None:
        main = ttk.Frame(self.content_frame, padding=24)
        main.pack(fill="both", expand=True)
        main.columnconfigure(1, weight=1)
        main.rowconfigure(8, weight=1)

        ###___Styles________________________________
        ### Available on all platform: alt, clam, classic, default
        # Create a style
        style = ttk.Style(main)
        # Set the theme with the theme_use method
        style.theme_use('default')  # put the theme name here, that you want to use
        ###_________________________________________

        ttk.Label(
            main,
            text="BIDS konvertering",
            font=("Segoe UI", 16, "bold"),
        ).grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 12))

        ttk.Label(
            main,
            text=(
                ("Välj en rootmapp, programmet kommer att leta efter all"
                "\nmappar som innehåller endast dicom filer")

            ),
            wraplength=1040,
            justify="left",
        ).grid(row=1, column=0, columnspan=3, sticky="w", pady=(0, 18))

        ttk.Label(main, text="Rotmapp:").grid(row=2, column=0, sticky="w", pady=8)
        ttk.Entry(main, textvariable=self.other_group_root_dir).grid(
            row=2, column=1, sticky="ew", padx=8
        )
        ttk.Button(
            main,
            text="Browse",
            command=self.select_other_group_root_dir,
        ).grid(row=2, column=2, sticky="ew")

        ttk.Label(main, text="Outputmapp:").grid(row=3, column=0, sticky="w", pady=8)
        ttk.Entry(main, textvariable=self.other_group_output_dir).grid(
            row=3, column=1, sticky="ew", padx=8
        )
        ttk.Button(
            main,
            text="Browse",
            command=self.select_other_group_output_dir,
        ).grid(row=3, column=2, sticky="ew")

        other_options_frame = ttk.LabelFrame(
            main,
            text="Valmöjligheter",
            padding=10
        )
        other_options_frame.grid(row=4,column=0,columnspan=3,sticky="ew",pady=(14,10))
        other_options_frame.rowconfigure(0,weight=1)
        other_options_frame.rowconfigure(1,weight=1)

        ttk.Label(
            other_options_frame,
            text="Parallella kärnor:"
        ).grid(row=0,column=0,sticky="w",pady=(10,15),padx=10)

        self.comob = ttk.Combobox(
            other_options_frame,
            textvariable=self.converter_workers_var,
            values=self.worker_options,
            width=8,
            state="readonly",
        )
        self.comob.grid(row=0, column=1, sticky="w", padx=10)

        ttk.Label(
            other_options_frame,
            textvariable=self.cpu_info_text,
        ).grid(row=0, column=2, columnspan=3, sticky="w", padx=10)

        ttk.Label(
            other_options_frame,
            text="Komprimera:"
        ).grid(row=1,column=0,sticky="w", pady=(10,0),padx=10)

        self.ifzipY = ttk.Radiobutton(
            other_options_frame,
            text = "Y",
            variable=self.YN,
            value=True
        ).grid(row=1,column=1,columnspan=2,padx=10,sticky="w")

        self.ifzipN = ttk.Radiobutton(
            other_options_frame,
            text = "N",
            variable=self.YN,
            value=False
        ).grid(row=1,column=2,columnspan=3,padx=10,sticky="w")

        ttk.Label(
            other_options_frame,
            text="'run-' i alla filnamn:"
        ).grid(row=2,column=0,sticky="w", pady=(10,0),padx=10)

        self.if_Run_Name_Y = ttk.Radiobutton(
            other_options_frame,
            text = "Y",
            variable=self.Run_in_name_bool,
            value=True
        ).grid(row=2,column=1,columnspan=2,padx=10,sticky="w")

        self.if_Run_Name_N = ttk.Radiobutton(
            other_options_frame,
            text = "N",
            variable=self.Run_in_name_bool,
            value=False
        ).grid(row=2,column=2,columnspan=3,padx=10,sticky="w")

        ttk.Label(
            other_options_frame,
            text="'Försäkra om filer är dcm'"
        ).grid(row=3,column=0,sticky="w", pady=(10,10),padx=10)

        self.if_Run_Name_Y = ttk.Radiobutton(
            other_options_frame,
            text = "Y",
            variable=self.guarantee_dcm,
            value=True
        ).grid(row=3,column=1,columnspan=2,padx=10,sticky="w")

        self.if_Run_Name_N = ttk.Radiobutton(
            other_options_frame,
            text = "N",
            variable=self.guarantee_dcm,
            value=False
        ).grid(row=3,column=2,columnspan=3,padx=10,sticky="w")

        ttk.Button(
            main,
            text="Run",
            command=self.run_bids_converter_starter,
        ).grid(row=5, column=0, columnspan=1, sticky="w", pady=(10, 0))

        ttk.Button(
            main,
            text="Break",
            command=self.break_bids_converter
        ).grid(row=5, column=1, columnspan=1, sticky="w", pady=(10, 0))

        ttk.Label(
            main,
            text="Pågående konvertering:"
        ).grid(row=6, column=0, sticky="w", pady=(10, 4))

        self.busy_progress_convert = ttk.Progressbar(
            main,
            orient="horizontal",
            mode="indeterminate"
        )
        self.busy_progress_convert.grid(row=7, column=0, columnspan=3, sticky="ew", padx=30, pady=(5,10))

        other_status_box = ttk.LabelFrame(main, text="Status", padding=10)
        other_status_box.grid(row=8, column=0, columnspan=3, sticky="nsew")
        other_status_box.columnconfigure(0, weight=1)
        other_status_box.rowconfigure(1, weight=1)

        ttk.Label(other_status_box, textvariable=self.other_status_text, wraplength=1040).grid(
            row=0, column=0, sticky="w", pady=(0, 10)
        )

        other_log_frame = ttk.Frame(other_status_box)
        other_log_frame.grid(row=1, column=0, sticky="nsew")
        other_log_frame.columnconfigure(0, weight=1)
        other_log_frame.rowconfigure(0, weight=1)

        self.other_log_widget = tk.Text(other_log_frame, wrap="word", height=24)
        self.other_log_widget.grid(row=0, column=0, sticky="nsew")

        other_log_scrollbar = ttk.Scrollbar(other_log_frame, orient="vertical", command=self.other_log_widget.yview)
        other_log_scrollbar.grid(row=0, column=1, sticky="ns", )

        self.other_log_widget.configure(yscrollcommand=other_log_scrollbar.set, state="disabled")
        other_log_frame.grid_propagate(True)
        

    def log(self, text: str) -> None:
        """
        Skriver en rad text i GUI:ts loggruta.
        """

        if not self._widget_exists(self.log_widget):
            return

        self.log_widget.configure(state="normal")
        self.log_widget.insert("end", text + "\n")
        self.log_widget.see("end")
        self.log_widget.configure(state="disabled")
        self.update_idletasks()

    def set_busy(self, busy: bool, current_text: Optional[str] = None) -> None:
        """
        Startar eller stoppar den indeterminate progressbaren.

        Används för att visa att något pågår även när exakt progress inte är känd.
        """

        if current_text is not None:
            self.current_item_text.set(current_text)

        if self._widget_exists(self.busy_progress):
            if busy:
                self.busy_progress.start(12)
            else:
                self.busy_progress.stop()

        self.update_idletasks()

    def set_progress(self, current_index: int, total: int) -> None:
        """
        Uppdaterar den deterministiska progressbaren baserat på antal färdiga sessioner.
        """

        if total <= 0:
            self.progress_value.set(0.0)
        else:
            self.progress_value.set((current_index / total) * 100.0)
        self.update_idletasks()

    def update_session_controls_state(self) -> None:
        """
        Aktiverar eller inaktiverar sessionrelaterade GUI-kontroller.

        Kontrollerna stängs av när programmet kör eller när inga färdiga sessioner
        finns att visa.
        """
        
        if self.current_view != "coregistrering_ui":
            self._safe_configure(self.session_combo, state="disabled")
            self._safe_configure(self.refresh_sessions_button, state="disabled")
            self._safe_configure(self.open_session_button, state="disabled")
            return

        has_sessions = bool(self.session_map)

        if self.is_running:
            self._safe_configure(self.session_combo, state="disabled")
            self._safe_configure(self.refresh_sessions_button, state="disabled")
            self._safe_configure(self.open_session_button, state="disabled")
            return

        self._safe_configure(self.refresh_sessions_button, state="normal")
        self._safe_configure(self.session_combo, state="readonly" if has_sessions else "disabled")
        can_open_in_fsl = has_sessions and not running_on_windows()
        self._safe_configure(
            self.open_session_button,
            state="normal" if can_open_in_fsl else "disabled"
        )

    def set_session_choices(self, items: List[Tuple[str, str]]) -> None:
        """
        Uppdaterar listan över valbara sessioner i comboboxen.
        """

        self.session_map = {label: path for label, path in items}
        values = list(self.session_map.keys())

        if self._widget_exists(self.session_combo):
            self.session_combo["values"] = values

        if values:
            self.session_choice.set(values[0])
        else:
            self.session_choice.set("")

        self.update_session_controls_state()

    def refresh_session_choices(self) -> None:
        """
        Läser in sessioner där färdiga coregistreringar finns.

        Resultatet visas i GUI:t så att användaren kan öppna T1 och coreg-filer i FSL.
        """

        if self.is_running or self.current_view != "coregistrering_ui":
            return

        root = self.root_dir.get().strip()
        if not root:
            self.set_session_choices([])
            messagebox.showwarning("Ingen rotmapp", "Välj rotmappen först.")
            return

        root_path = Path(root)
        if not root_path.exists():
            self.set_session_choices([])
            messagebox.showerror("Error", f"Rotmappen finns inte:\n{root}")
            return

        if not root_path.is_dir():
            self.set_session_choices([])
            messagebox.showerror("Error", f"Det här är inte en mapp:\n{root}")
            return

        all_sessions = find_registered_subject_session_dirs(root_path)
        items = [
            (f"{subject} / {session}", str(session_dir))
            for subject, session, session_dir in all_sessions
        ]

        self.set_session_choices(items)

        self.status_text.set(
            f"Hittade {len(items)} session(er) med färdig coregistrering. "
            "Välj en session och öppna T1 + coreg-filer i FSL."
        )

        self.log(f"Sessioner med färdig coregistrering i: {root_path}")
        for label, path in items:
            self.log(f"  {label}: {path}")
        self.log("")

    def select_fsl_manually(self) -> None:
        """
        Låter användaren välja eller ändra sparad FSL/flirt-sökväg manuellt. Låter användaren välja eller ändra sparad FSL/flirt-sökväg manuellt.
        """

        if running_on_windows():
            self.status_text.set("FSL är avstängt när programmet körs nativt i Windows.")
            messagebox.showinfo(
                "FSL avstängt på Windows",
                WINDOWS_FSL_DISABLED_MESSAGE,
                parent=self,
            )
            return

        if self.is_running:
            messagebox.showwarning(
                "Körning pågår",
                "Ändra inte FSL-sökväg medan programmet kör."
            )
            return

        try:
            selected = prompt_user_for_fsl_location(self)

            if selected is None:
                return

            configure_fsl_from_flirt(selected)
            remember_fsl_flirt_path(selected)

            self.status_text.set(f"FSL-sökväg sparad: {selected}")
            self.log(f"FSL-sökväg sparad: {selected}")

            messagebox.showinfo(
                "FSL sparat",
                "FSL-sökvägen har sparats för framtida körningar.\n\n"
                f"flirt:\n{selected}"
            )

        except Exception as exc:
            self.status_text.set(f"Fel vid val av FSL: {exc}")
            messagebox.showerror(
                "FSL error",
                f"{exc}\n\nDetaljer:\n{traceback.format_exc()}"
            )

    def validate_root_dir_for_similarity_metrics(self) -> Path:
        """
        Kontrollerar att rotmappen och nödvändiga paket finns för similarity metrics.
        """

        missing = []

        if nib is None:
            missing.append("nibabel")
        if np is None:
            missing.append("numpy")
        if Workbook is None:
            missing.append("openpyxl")

        if missing:
            raise RuntimeError(
                "Följande Python-paket saknas: "
                + ", ".join(missing)
                + ". Installera med: pip install "
                + " ".join(missing)
            )

        root = self.root_dir.get().strip()

        if not root:
            raise ValueError("Välj rotmappen först.")

        root_path = Path(root)

        if not root_path.exists():
            raise FileNotFoundError(root)

        if not root_path.is_dir():
            raise NotADirectoryError(root)

        return root_path

    def create_similarity_metrics_from_menu(self) -> None:
        """
        Startar skapandet av similarity metrics-Excel från menyn.

        Arbetet körs i en separat tråd så att GUI:t inte fryser.
        """

        if self.is_running:
            messagebox.showwarning(
                "Körning pågår",
                "Vänta tills nuvarande körning är klar innan du skapar similarity metrics."
            )
            return

        try:
            root_path = self.validate_root_dir_for_similarity_metrics()
        except Exception as exc:
            self.status_text.set(f"Fel: {exc}")
            messagebox.showerror("Error", f"{exc}\n\nDetaljer:\n{traceback.format_exc()}")
            return

        self.set_controls_running(True)
        self.set_busy(True, "Skapar similarity metrics Excel...")
        self.status_text.set("Skapar similarity metrics Excel från färdiga coregistreringar...")

        self.worker_thread = threading.Thread(
            target=self._worker_create_similarity_metrics,
            args=(root_path,),
            daemon=True,
        )
        self.worker_thread.start()

    def _worker_create_similarity_metrics(self, root_path: Path) -> None:
        """
        Worker-funktion som skapar similarity metrics-Excel i bakgrunden.
        """

        try:
            summary_root = root_path / "derivatives" / "registration"
            summary_root.mkdir(parents=True, exist_ok=True)

            self.ui_queue.put(("log", "Skapar similarity metrics Excel från befintliga coregistreringar..."))
            self.ui_queue.put(("log", f"Rotmapp: {root_path}"))

            rows = collect_existing_similarity_rows_from_outputs(root_path)

            similarity_metrics_path = summary_root / SIMILARITY_METRICS_FILENAME

            write_similarity_metrics_xlsx(
                path=str(similarity_metrics_path),
                rows=rows,
            )

            self.ui_queue.put((
                "metrics_done",
                {
                    "similarity_metrics_path": str(similarity_metrics_path),
                    "row_count": len(rows),
                },
            ))

        except Exception as exc:
            self.ui_queue.put((
                "metrics_error",
                {
                    "message": str(exc),
                    "traceback": traceback.format_exc(),
                },
            ))

    def open_similarity_metrics_from_menu(self) -> None:
        """
        Öppnar similarity metrics-Excel.

        Om filen inte finns får användaren möjlighet att skapa den först.
        """

        try:
            root_path = self.validate_root_dir_for_similarity_metrics()
            similarity_metrics_path = (
                root_path
                / "derivatives"
                / "registration"
                / SIMILARITY_METRICS_FILENAME
            )

            if not similarity_metrics_path.exists():
                answer = messagebox.askyesno(
                    "Similarity metrics saknas",
                    "Excel-filen med similarity metrics finns inte ännu.\n\n"
                    "Vill du skapa den nu?"
                )

                if answer:
                    self.create_similarity_metrics_from_menu()

                return

            open_file_with_default_app(similarity_metrics_path)

            self.status_text.set(
                f"Öppnade similarity metrics Excel: {similarity_metrics_path}"
            )
            self.log(f"Öppnade similarity metrics Excel: {similarity_metrics_path}")

        except Exception as exc:
            self.status_text.set(f"Fel: {exc}")
            messagebox.showerror("Error", f"{exc}\n\nDetaljer:\n{traceback.format_exc()}")

    def validate_checkerboard_selection(self) -> Tuple[Path, Path, str]:
        """
        Kontrollerar att en giltig session är vald för checkerboard-QC.
        """

        if nib is None or np is None:
            raise RuntimeError(
                "Checkerboard QC kräver nibabel och numpy. "
                "Installera med: pip install nibabel numpy"
            )

        root = self.root_dir.get().strip()

        if not root:
            raise ValueError("Välj rotmappen först.")

        root_path = Path(root)

        if not root_path.exists():
            raise FileNotFoundError(root)

        if not root_path.is_dir():
            raise NotADirectoryError(root)

        selection = self.session_choice.get().strip()
        session_path_str = self.session_map.get(selection, "")

        if not session_path_str:
            raise ValueError(
                "Välj en färdig session först. "
                "Tryck på 'Läs in färdiga sessioner' om listan är tom."
            )

        session_path = Path(session_path_str)

        if not session_path.exists():
            raise FileNotFoundError(session_path)

        return root_path, session_path, selection

    def create_checkerboard_from_menu(self) -> None:
        """
        Startar skapande av checkerboard-QC för vald session.
        """

        if self.is_running:
            messagebox.showwarning(
                "Körning pågår",
                "Vänta tills nuvarande körning är klar innan du skapar checkerboard QC."
            )
            return

        try:
            root_path, session_path, selection = self.validate_checkerboard_selection()
        except Exception as exc:
            self.status_text.set(f"Fel: {exc}")
            messagebox.showerror("Error", f"{exc}\n\nDetaljer:\n{traceback.format_exc()}")
            return

        self.set_controls_running(True)
        self.set_busy(True, f"Skapar checkerboard QC för {selection}...")
        self.status_text.set(f"Skapar checkerboard QC för {selection}...")

        self.worker_thread = threading.Thread(
            target=self._worker_create_checkerboard,
            args=(root_path, session_path, selection),
            daemon=True,
        )
        self.worker_thread.start()

    def create_checkerboard_all_from_menu(self) -> None:
        """
        Startar skapande av checkerboard-QC för alla färdiga sessioner.
        """

        if self.is_running:
            messagebox.showwarning(
                "Körning pågår",
                "Vänta tills nuvarande körning är klar innan du skapar checkerboard QC."
            )
            return

        root = self.root_dir.get().strip()

        if not root:
            messagebox.showwarning("Ingen rotmapp", "Välj rotmappen först.")
            return

        root_path = Path(root)

        if not root_path.exists():
            messagebox.showerror("Error", f"Rotmappen finns inte:\n{root_path}")
            return

        if not root_path.is_dir():
            messagebox.showerror("Error", f"Det här är inte en mapp:\n{root_path}")
            return

        if nib is None or np is None:
            messagebox.showerror(
                "Saknade paket",
                "Checkerboard QC kräver nibabel och numpy.\n\n"
                "Installera med:\n"
                "pip install nibabel numpy"
            )
            return

        self.set_controls_running(True)
        self.set_busy(True, "Skapar checkerboard QC för alla färdiga sessioner...")
        self.status_text.set("Skapar checkerboard QC för alla färdiga sessioner...")

        self.worker_thread = threading.Thread(
            target=self._worker_create_checkerboard_all,
            args=(root_path,),
            daemon=True,
        )
        self.worker_thread.start()

    def _worker_create_checkerboard_all(self, root_path: Path) -> None:
        """
        Worker-funktion som skapar checkerboard-QC för alla registrerade sessioner.
        """

        try:
            self.ui_queue.put(("log", "Skapar checkerboard QC för alla färdiga sessioner..."))
            self.ui_queue.put(("log", f"Rotmapp: {root_path}"))

            results = create_checkerboards_for_all_registered_sessions(
                root_dir=root_path,
                block_size=CHECKERBOARD_BLOCK_SIZE,
            )

            processed_sessions = sum(
                1 for item in results
                if item.get("status") == "processed"
            )

            failed_sessions = sum(
                1 for item in results
                if item.get("status") == "failed"
            )

            checkerboard_count = sum(
                int(item.get("checkerboard_count") or 0)
                for item in results
            )

            self.ui_queue.put((
                "checkerboard_all_done",
                {
                    "results": results,
                    "processed_sessions": processed_sessions,
                    "failed_sessions": failed_sessions,
                    "checkerboard_count": checkerboard_count,
                },
            ))

        except Exception as exc:
            self.ui_queue.put((
                "checkerboard_all_error",
                {
                    "message": str(exc),
                    "traceback": traceback.format_exc(),
                },
            ))

    def _worker_create_checkerboard(
        self,
        root_path: Path,
        session_path: Path,
        selection: str,
    ) -> None:
        """
        Worker-funktion som skapar checkerboard-QC för en vald session.
        """

        try:
            self.ui_queue.put(("log", f"Skapar checkerboard QC för: {selection}"))
            self.ui_queue.put(("log", f"Session: {session_path}"))

            results = create_checkerboards_for_session(
                root_dir=root_path,
                session_dir=session_path,
                block_size=CHECKERBOARD_BLOCK_SIZE,
            )

            reference_file = results[0]["reference_file"] if results else None
            checkerboard_files = [
                item["checkerboard_file"]
                for item in results
                if item.get("checkerboard_file")
            ]

            self.ui_queue.put((
                "checkerboard_done",
                {
                    "selection": selection,
                    "reference_file": reference_file,
                    "checkerboard_files": checkerboard_files,
                    "items": results,
                },
            ))

        except Exception as exc:
            self.ui_queue.put((
                "checkerboard_error",
                {
                    "message": str(exc),
                    "traceback": traceback.format_exc(),
                },
            ))

    def open_checkerboard_from_menu(self) -> None:
        """
        Öppnar befintliga checkerboard-QC-filer för vald session i FSL.

        Om filerna saknas får användaren möjlighet att skapa dem först.
        På Windows är öppning i FSL avstängd.
        """

        if running_on_windows():
            self.status_text.set(
                "Kan inte öppna checkerboard-QC i FSL när programmet körs nativt i Windows."
            )
            messagebox.showinfo(
                "Öppning i FSL avstängd",
                WINDOWS_FSL_DISABLED_MESSAGE,
                parent=self,
            )
            return

        try:
            root_path, session_path, selection = self.validate_checkerboard_selection()

            checkerboard_files = find_checkerboard_qc_files_for_session(
                root_dir=root_path,
                session_dir=session_path,
            )

            if not checkerboard_files:
                answer = messagebox.askyesno(
                    "Checkerboard QC saknas",
                    "Det finns inga checkerboard-QC-filer för vald session ännu.\n\n"
                    "Vill du skapa dem nu?"
                )

                if answer:
                    self.create_checkerboard_from_menu()

                return

            viewer = find_fsl_viewer_executable()
            if viewer is None:
                raise RuntimeError(
                    "Hittade ingen FSL-visare. Testade: fsleyes, fslview_deprecated, fslview."
                )

            reference_file = find_reference_t1_for_session(root_path, session_path)

            files_to_open = []
            if reference_file is not None:
                files_to_open.append(str(reference_file))

            files_to_open.extend(str(p) for p in checkerboard_files)

            subprocess.Popen([viewer] + files_to_open)

            self.status_text.set(
                f"Öppnade {len(files_to_open)} fil(er) med checkerboard QC för {selection}."
            )

            self.log(f"Öppnade checkerboard QC i FSL: {selection}")
            for path in files_to_open:
                self.log(f"  {path}")
            self.log("")

        except Exception as exc:
            self.status_text.set(f"Fel vid öppning av checkerboard QC: {exc}")
            messagebox.showerror(
                "Error",
                f"{exc}\n\nDetaljer:\n{traceback.format_exc()}"
            )
        

    def open_selected_session_in_fsl(self) -> None:
        """
        Öppnar T1-referensen och coregistrerade filer för vald session i FSL.
        """

        if running_on_windows():
            self.status_text.set("Kan inte öppna NIfTI-bilder i FSL när programmet körs nativt i Windows.")
            messagebox.showinfo(
                "Öppning i FSL avstängd",
                WINDOWS_FSL_DISABLED_MESSAGE,
                parent=self,
            )
            return

        selection = self.session_choice.get().strip()
        session_path_str = self.session_map.get(selection, "")

        if not session_path_str:
            messagebox.showwarning("Ingen session vald", "Välj en session först.")
            return

        try:
            viewer = find_fsl_viewer_executable()
            if viewer is None:
                raise RuntimeError(
                    "Hittade ingen FSL-visare. Testade: fsleyes, fslview_deprecated, fslview."
                )

            root_path = Path(self.root_dir.get().strip())
            session_path = Path(session_path_str)

            nifti_files = collect_t1_space_files_for_fsl(root_path, session_path)
            coreg_files = [p for p in nifti_files if is_coregistered_to_t1_file(p)]

            if not coreg_files:
                raise FileNotFoundError(
                    "Hittade inga färdiga coregistrerade filer för den här sessionen."
                )

            subprocess.Popen([viewer] + [str(p) for p in nifti_files])

            self.status_text.set(
                f"Öppnade {len(nifti_files)} fil(er) i FSL för sessionen {selection}."
            )
            self.log(f"Öppnade T1 + coreg i FSL: {selection}")
            for path in nifti_files:
                self.log(f"  {path}")
            self.log("")

        except Exception as exc:
            self.status_text.set(f"Fel vid öppning i FSL: {exc}")
            messagebox.showerror("Error", f"{exc}\n\nDetaljer:\n{traceback.format_exc()}")

    def set_controls_running(self, running: bool) -> None:
        """
        Aktiverar eller inaktiverar GUI-kontroller beroende på om en körning pågår.
        """

        self.is_running = running
        self._safe_configure(self.run_button, state="disabled" if running else "normal")
        self._safe_configure(self.clear_button, state="disabled" if running else "normal")
        self._safe_configure(self.use_contrast_check, state="disabled" if running else "normal")
        self._safe_configure(self.run_t2_check, state="disabled" if running else "normal")
        self._safe_configure(self.run_diffusion_check, state="disabled" if running else "normal")
        self._safe_configure(self.run_perfusion_check, state="disabled" if running else "normal")
        self._safe_configure(self.validate_bids_check, state="disabled" if running else "normal")
        self._safe_configure(self.diffusion_cost_combo, state="disabled" if running else "readonly")
        self._safe_configure(self.perfusion_cost_combo, state="disabled" if running else "readonly")
        self._safe_configure(self.max_workers_combo, state="disabled" if running else "readonly")
        self.update_session_controls_state()

    def select_root_dir(self) -> None:
        """
        Öppnar en dialog där användaren väljer datasetets rotmapp.
        """

        if self.is_running:
            return

        path = filedialog.askdirectory(
            title = "Välj rotmappen som innehåller sub-*/ses-*",
            )

        if path:
            self.root_dir.set(path)
            if self.current_view == "coregistrering_ui":
                self.refresh_session_choices()

    def clear_fields(self) -> None:
        """
        Återställer alla GUI-fält, statusrader, progressbars och loggen.
        """

        if self.is_running:
            return

        self.root_dir.set("")
        self.run_t2_var.set(False)
        self.run_diffusion_var.set(False)
        self.run_perfusion_var.set(False)
        self.use_contrast_var.set(False)
        self.validate_bids_var.set(VALIDATE_BIDS_BEFORE_RUN_DEFAULT)
        self.diffusion_cost_var.set(DEFAULT_DIFFUSION_COST)
        self.perfusion_cost_var.set(DEFAULT_PERFUSION_COST)
        self.other_group_root_dir.set("")
        self.other_group_output_dir.set("")
        self.max_workers_var.set(str(self.recommended_cpu_workers))
        self.converter_workers_var.set(str(self.recommended_cpu_workers))

        self.status_text.set("Välj rotmappen som innehåller sub-*/ses-* och tryck Run.")
        self.current_item_text.set("Ingen körning aktiv.")
        self.progress_value.set(0.0)
        

        if self._widget_exists(self.busy_progress):
            self.busy_progress.stop()

        self.set_session_choices([])

        if self._widget_exists(self.log_widget):
            self.log_widget.configure(state="normal")
            self.log_widget.delete("1.0", "end")
            self.log_widget.configure(state="disabled")

    def validate_inputs(self) -> Path:
        """
        Validerar användarens input innan batch-körningen startar.

        Funktionen kontrollerar Python-beroenden, FSL och att rotmappen finns.
        """

        if running_on_windows():
            raise RuntimeError(WINDOWS_FSL_DISABLED_MESSAGE)

        ensure_python_dependencies()
        flirt_path = ensure_fsl_available_gui(self)
        self.log(f"FSL flirt: {flirt_path}")

        root = self.root_dir.get().strip()
        if not root:
            raise ValueError("Välj rotmappen som innehåller sub-*/ses-*.")

        root_path = Path(root)
        if not root_path.exists():
            raise FileNotFoundError(root)
        if not root_path.is_dir():
            raise NotADirectoryError(root)

        return root_path
    def process_ui_queue(self) -> None:
        """
        Läser statusmeddelanden från worker-trådar och uppdaterar GUI:t.

        Tkinter ska bara uppdateras från huvudtråden. Därför skickar
        bakgrundstrådar meddelanden till self.ui_queue, och denna funktion
        hämtar dem regelbundet med hjälp av self.after().
        """

        try:
            while True:
                event_type, payload = self.ui_queue.get_nowait()

                if event_type == "log":
                    self.log(payload)

                elif event_type == "status":
                    self.status_text.set(payload)

                elif event_type == "current":
                    self.current_item_text.set(payload)

                elif event_type == "progress":
                    current_index, total = payload
                    self.set_progress(current_index, total)

                elif event_type == "busy":
                    busy, text = payload
                    self.set_busy(busy, text)

                elif event_type == "done":
                    self.set_controls_running(False)
                    self.set_busy(False, "Ingen körning aktiv.")

                    summary_path = payload["summary_path"]
                    similarity_metrics_path = payload.get("similarity_metrics_path")
                    processed_count = payload["processed_count"]
                    skipped_count = payload["skipped_count"]
                    failed_count = payload["failed_count"]
                    elapsed_time = payload.get("elapsed_time", "okänd tid")

                    self.status_text.set(
                        "Klart. "
                        f"Total tid: {elapsed_time}. "
                        f"Bearbetade registreringar: {processed_count}, "
                        f"hoppade över befintliga: {skipped_count}, "
                        f"misslyckade sessioner: {failed_count}. "
                        f"Sammanfattning: {summary_path}. "
                        f"Similarity metrics: {similarity_metrics_path}"
                    )

                    if self.current_view == "coregistrering_ui":
                        self.refresh_session_choices()

                    if self.close_when_done:
                        self.destroy()
                    else:
                        messagebox.showinfo(
                            "Klart",
                            "Batch-coregistrering klar.\n\n"
                            f"Total tid: {elapsed_time}\n"
                            f"Bearbetade registreringar: {processed_count}\n"
                            f"Hoppade över befintliga: {skipped_count}\n"
                            f"Misslyckade sessioner: {failed_count}\n\n"
                            f"Sammanfattning:\n{summary_path}\n\n"
                            f"Similarity metrics:\n{similarity_metrics_path}"
                        )

                elif event_type == "metrics_done":
                    self.set_controls_running(False)
                    self.set_busy(False, "Ingen körning aktiv.")

                    similarity_metrics_path = payload["similarity_metrics_path"]
                    row_count = payload["row_count"]

                    self.status_text.set(
                        f"Similarity metrics Excel skapad med {row_count} rad(er): "
                        f"{similarity_metrics_path}"
                    )

                    self.log(
                        f"Similarity metrics Excel skapad med {row_count} rad(er): "
                        f"{similarity_metrics_path}"
                    )
                    self.log("")

                    messagebox.showinfo(
                        "Similarity metrics klar",
                        "Similarity metrics Excel skapad.\n\n"
                        f"Antal rader: {row_count}\n\n"
                        f"Fil:\n{similarity_metrics_path}"
                    )

                elif event_type == "checkerboard_done":
                    self.set_controls_running(False)
                    self.set_busy(False, "Ingen körning aktiv.")

                    selection = payload["selection"]
                    reference_file = payload.get("reference_file")
                    checkerboard_files = payload.get("checkerboard_files", [])
                    items = payload.get("items", [])

                    self.status_text.set(
                        f"Checkerboard QC skapad för {selection}. "
                        f"Antal filer: {len(checkerboard_files)}."
                    )

                    self.log(f"Checkerboard QC skapad för {selection}:")
                    for item in items:
                        self.log(f"  Reference: {item.get('reference_file')}")
                        self.log(f"  Registered: {item.get('registered_file')}")
                        self.log(f"  Checkerboard: {item.get('checkerboard_file')}")
                        self.log(f"  JSON: {item.get('checkerboard_json')}")
                        self.log("")
                    self.log("")

                    if running_on_windows():
                        self.status_text.set(
                            "Checkerboard QC skapades, men öppning i FSL är avstängd på Windows."
                        )
                        messagebox.showinfo(
                            "Checkerboard QC klar",
                            "Checkerboard QC skapades.\n\n"
                            f"Antal filer: {len(checkerboard_files)}\n\n"
                            "Programmet körs nativt på Windows, så bilderna öppnas inte i FSL.",
                            parent=self,
                        )
                    else:
                        answer = messagebox.askyesno(
                            "Checkerboard QC klar",
                            "Checkerboard QC skapad.\n\n"
                            f"Antal filer: {len(checkerboard_files)}\n\n"
                            "Vill du öppna dem i FSL nu?"
                        )

                        if answer:
                            try:
                                viewer = find_fsl_viewer_executable()
                                if viewer is None:
                                    raise RuntimeError(
                                        "Hittade ingen FSL-visare. Testade: fsleyes, "
                                        "fslview_deprecated, fslview."
                                    )

                                files_to_open = []

                                if reference_file:
                                    files_to_open.append(str(reference_file))

                                files_to_open.extend(str(p) for p in checkerboard_files)

                                subprocess.Popen([viewer] + files_to_open)

                                self.status_text.set(
                                    f"Öppnade checkerboard QC i FSL för {selection}."
                                )

                            except Exception as exc:
                                self.status_text.set(
                                    f"Checkerboard skapades, men kunde inte öppnas i FSL: {exc}"
                                )
                                messagebox.showerror(
                                    "FSL error",
                                    f"{exc}\n\nDetaljer:\n{traceback.format_exc()}"
                                )

                elif event_type == "checkerboard_all_done":
                    self.set_controls_running(False)
                    self.set_busy(False, "Ingen körning aktiv.")

                    results = payload["results"]
                    processed_sessions = payload["processed_sessions"]
                    failed_sessions = payload["failed_sessions"]
                    checkerboard_count = payload["checkerboard_count"]

                    self.status_text.set(
                        "Checkerboard QC klar. "
                        f"Sessioner klara: {processed_sessions}, "
                        f"misslyckade: {failed_sessions}, "
                        f"checkerboard-filer: {checkerboard_count}."
                    )

                    self.log("Checkerboard QC klar för alla färdiga sessioner:")
                    self.log(f"  Sessioner klara: {processed_sessions}")
                    self.log(f"  Misslyckade sessioner: {failed_sessions}")
                    self.log(f"  Checkerboard-filer: {checkerboard_count}")
                    self.log("")

                    for item in results:
                        self.log(
                            f"  {item.get('subject')} / {item.get('session')}: "
                            f"{item.get('status')} "
                            f"({item.get('checkerboard_count')} filer)"
                        )

                        if item.get("status") == "failed":
                            self.log(f"    FEL: {item.get('error')}")

                    self.log("")

                    messagebox.showinfo(
                        "Checkerboard QC klar",
                        "Checkerboard QC skapad för alla färdiga sessioner.\n\n"
                        f"Sessioner klara: {processed_sessions}\n"
                        f"Misslyckade sessioner: {failed_sessions}\n"
                        f"Checkerboard-filer: {checkerboard_count}\n\n"
                        "Filerna ligger i:\n"
                        "derivatives/registration/checkerboard_qc/sub-*/ses-*"
                    )


                elif event_type == "checkerboard_all_error":
                    self.set_controls_running(False)
                    self.set_busy(False, "Ingen körning aktiv.")

                    error_message = payload["message"]
                    error_traceback = payload["traceback"]

                    self.status_text.set(f"Fel vid checkerboard QC: {error_message}")

                    messagebox.showerror(
                        "Checkerboard QC error",
                        f"{error_message}\n\nDetaljer:\n{error_traceback}"
                    )

                elif event_type == "checkerboard_error":
                    self.set_controls_running(False)
                    self.set_busy(False, "Ingen körning aktiv.")

                    error_message = payload["message"]
                    error_traceback = payload["traceback"]

                    self.status_text.set(f"Fel vid checkerboard QC: {error_message}")

                    messagebox.showerror(
                        "Checkerboard QC error",
                        f"{error_message}\n\nDetaljer:\n{error_traceback}"
                    )

                elif event_type == "metrics_error":
                    self.set_controls_running(False)
                    self.set_busy(False, "Ingen körning aktiv.")

                    error_message = payload["message"]
                    error_traceback = payload["traceback"]

                    self.status_text.set(f"Fel vid similarity metrics: {error_message}")

                    messagebox.showerror(
                        "Similarity metrics error",
                        f"{error_message}\n\nDetaljer:\n{error_traceback}"
                    )

                elif event_type == "error":
                    self.set_controls_running(False)
                    self.set_busy(False, "Ingen körning aktiv.")

                    error_message = payload["message"]
                    error_traceback = payload["traceback"]

                    self.status_text.set(f"Fel: {error_message}")

                    if self.close_when_done:
                        self.destroy()
                    else:
                        messagebox.showerror("Error", f"{error_message}\n\nDetaljer:\n{error_traceback}")

        except queue.Empty:
            pass

        self.after(150, self.process_ui_queue)

    def run_processing(self) -> None:
        """
        Startar batch-coregistreringen från GUI:t.

        Funktionen validerar först användarens input, läser inställningarna från
        gränssnittet och startar sedan en separat worker-tråd där själva
        bearbetningen sker.
        """

        if self.is_running:
            return

        try:
            root_path = self.validate_inputs()
        except Exception as exc:
            self.status_text.set(f"Fel: {exc}")
            messagebox.showerror("Error", f"{exc}\n\nDetaljer:\n{traceback.format_exc()}")
            return

        run_t2 = bool(self.run_t2_var.get())
        run_diffusion = bool(self.run_diffusion_var.get())
        run_perfusion = bool(self.run_perfusion_var.get())
        validate_bids = bool(self.validate_bids_var.get())
        use_contrast = bool(self.use_contrast_var.get())
        max_workers = int(self.max_workers_var.get())
        diffusion_cost = self.diffusion_cost_var.get().strip()
        perfusion_cost = self.perfusion_cost_var.get().strip()

        self.cancel_event.clear()
        self.close_when_done = False
        self.runtime_control = RuntimeControl(cancel_event=self.cancel_event)
        self.set_controls_running(True)
        self.progress_value.set(0.0)
        self.set_busy(True, "Startar batch-körning...")

        self.worker_thread = threading.Thread(
            target=self._worker_run_processing,
            args=(
                root_path,
                run_t2,
                run_diffusion,
                run_perfusion,
                use_contrast,
                validate_bids,
                max_workers,
                diffusion_cost,
                perfusion_cost,
            ),
            daemon=True,
        )
        self.worker_thread.start()

    def _worker_run_processing(
        self,
        root_path: Path,
        run_t2: bool,
        run_diffusion: bool,
        run_perfusion: bool,
        use_contrast: bool,
        validate_bids: bool,
        max_workers: int,
        diffusion_cost: str,
        perfusion_cost: str,
    ) -> None:
        """
        Kör hela batch-processen i bakgrunden.

        Funktionen hittar alla sessioner, kör eventuell BIDS-validering,
        processar sessionerna parallellt och sparar sedan sammanfattning,
        similarity metrics och logginformation.
        """
        try:
            batch_started_at = datetime.now()
            batch_start_perf = time.perf_counter()

            all_sessions = find_subject_session_dirs(root_path)
            if not all_sessions:
                raise RuntimeError("Hittade inga mappar som följer mönstret sub-*/ses-*.")

            summary_root = root_path / "derivatives" / "registration"
            summary_root.mkdir(parents=True, exist_ok=True)

            bids_validation_result: Dict[str, Any] = {
                "enabled": bool(validate_bids),
                "valid": None,
                "report_file": None,
            }

            if validate_bids:
                self.ui_queue.put(("status", "Kör BIDS-validering..."))
                self.ui_queue.put(("busy", (True, "Validerar BIDS-dataset innan coregistrering...")))
                self.ui_queue.put(("log", "Kör BIDS Validator endast på derivatives/registration..."))
                #self.ui_queue.put(("log", f"Validerar mapp: {registration_root}"))
                self.ui_queue.put(("log", f"Prune derivatives/sourcedata: {BIDS_VALIDATOR_PRUNE_DERIVATIVES}"))

                registration_root = root_path / "derivatives" / "registration"
                registration_root.mkdir(parents=True, exist_ok=True)

                ensure_derivative_dataset_description(registration_root)

                removed_gradients = remove_coreg_dwi_gradient_sidecars(registration_root)

                self.ui_queue.put((
                    "log",
                    f"Tog bort {removed_gradients} gamla .bval/.bvec-filer för coreg-DWI derivatives."
                ))

                # Viktigt: gammal work-mapp gör att BIDS Validator stoppar.
                remove_registration_work_before_bids_validation(registration_root)

                # Viktigt: gamla coreg-JSON-filer saknar ibland SpatialReference.
                updated_sidecars = ensure_required_derivative_sidecar_keys(registration_root)

                self.ui_queue.put((
                    "log",
                    f"Uppdaterade required derivative keys i {updated_sidecars} gamla coreg-JSON-filer."
                ))

                validation = run_bids_validator(
                    root_dir=registration_root,
                    runtime=self.runtime_control,
                    prune_derivatives=False,
                )

                bids_report_path = summary_root / BIDS_VALIDATOR_REPORT_FILENAME

                write_bids_validator_report(
                    path=bids_report_path,
                    root_dir=root_path,
                    validation_result=validation,
                )

                bids_validation_result = {
                    "enabled": True,
                    "valid": bool(validation["valid"]),
                    "returncode": validation["returncode"],
                    "command": validation["command"],
                    "report_file": str(bids_report_path),
                }

                self.ui_queue.put(("log", f"BIDS Validator returncode: {validation['returncode']}"))
                self.ui_queue.put(("log", f"BIDS Validator rapport: {bids_report_path}"))

                if not validation["valid"]:
                    output_preview = validation.get("output", "")
                    if len(output_preview) > 4000:
                        output_preview = output_preview[:4000] + "\n\n[truncated]"

                    raise RuntimeError(
                        "BIDS-validering misslyckades. Coregistreringen stoppades.\n\n"
                        f"Rapport:\n{bids_report_path}\n\n"
                        f"Validator-output:\n{output_preview}"
                    )

                self.ui_queue.put(("log", "BIDS-validering OK. Fortsätter med coregistrering."))
                self.ui_queue.put(("log", ""))
            else:
                self.ui_queue.put(("log", "BIDS-validering är avstängd för denna körning."))
                self.ui_queue.put(("log", ""))

            work_root = root_path / "derivatives" / "registration_work"
            work_root.mkdir(parents=True, exist_ok=True)

            self.ui_queue.put(("status", f"Hittade {len(all_sessions)} session(er). Kör coregistrering..."))
            self.ui_queue.put(("busy", (True, f"Kör coregistrering för {len(all_sessions)} session(er)...")))
            self.ui_queue.put(("log", f"Rotmapp: {root_path}"))
            self.ui_queue.put(("log", f"Starttid: {batch_started_at.isoformat(timespec='seconds')}"))
            self.ui_queue.put(("log", f"Default cost (PET/T2): {DEFAULT_COST}"))
            self.ui_queue.put(("log", f"Diffusion cost: {diffusion_cost}"))
            self.ui_queue.put(("log", f"Perfusion cost: {perfusion_cost}"))
            self.ui_queue.put(("log", f"Default DOF: {DEFAULT_DOF}"))
            self.ui_queue.put(("log", f"Använd kontrast: {use_contrast}"))
            self.ui_queue.put(("log", f"Kör T2 -> T1: {run_t2}"))
            self.ui_queue.put(("log", f"Kör diffusion -> T1: {run_diffusion}"))
            self.ui_queue.put(("log", f"Kör perfusion -> T1: {run_perfusion}"))
            self.ui_queue.put(("log", f"Parallella sessioner: {max_workers}"))
            self.ui_queue.put(("log", ""))

            processed_count = 0
            skipped_count = 0
            failed_count = 0
            session_results: List[Dict[str, Any]] = []

            total_sessions = len(all_sessions)
            completed_sessions = 0

            self.ui_queue.put(("progress", (0, total_sessions)))

            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                self.executor = executor

                def run_one_session(item):
                    """
                    Kör processering för en enskild session i trådpoolen.

                    Funktionen skickar status till GUI-kön och anropar process_session().
                    """

                    subject, session, session_dir = item

                    self.ui_queue.put((
                        "current",
                        f"Kör nu: {subject}/{session} ({session_dir})"
                    ))
                    self.ui_queue.put((
                        "log",
                        f"START {subject}/{session}: {datetime.now().isoformat(timespec='seconds')}"
                    ))

                    return process_session(
                        root_path,
                        session_dir,
                        DEFAULT_COST,
                        diffusion_cost,
                        perfusion_cost,
                        DEFAULT_DOF,
                        work_root,
                        run_t2,
                        run_diffusion,
                        run_perfusion,
                        use_contrast,
                        self.runtime_control,
                    )

                future_map = {
                    executor.submit(run_one_session, item): item[2]
                    for item in all_sessions
                }

                for future in as_completed(future_map):
                    session_dir = future_map[future]
                    completed_sessions += 1

                    self.ui_queue.put(("status", f"Klart med session {completed_sessions}/{total_sessions}: {session_dir}"))
                    self.ui_queue.put(("busy", (True, f"Bearbetat {completed_sessions} av {total_sessions} sessioner. Senast klar: {session_dir}")))
                    self.ui_queue.put(("log", f"[{completed_sessions}/{total_sessions}] {session_dir}"))

                    try:
                        result = future.result()
                        session_results.append(result)

                        session_elapsed_time = result.get("elapsed_time", "okänd tid")

                        pet_status = result["pet"]["status"] if result.get("pet") else None
                        t2_status = result["t2"]["status"] if result.get("t2") else None
                        diffusion_status = result["diffusion"]["status"] if result.get("diffusion") else None
                        perfusion_status = result["perfusion"]["status"] if result.get("perfusion") else None

                        for status in (pet_status, t2_status, diffusion_status, perfusion_status):
                            if status == "processed":
                                processed_count += 1
                            elif status == "skipped_existing":
                                skipped_count += 1

                        self.ui_queue.put(("log", f"  T1: {result['t1_reference']}"))
                        self.ui_queue.put(("log", f"  Sessionstid: {session_elapsed_time}"))
                        self.ui_queue.put(("log", f"  PET: {pet_status}"))
                        self.ui_queue.put(("log", f"  PET output: {result['pet']['output_file']}"))
                        self.ui_queue.put(("log", f"  PET JSON: {result['pet']['output_json']}"))

                        if result.get("t2") is not None:
                            self.ui_queue.put(("log", f"  T2: {t2_status}"))
                            self.ui_queue.put(("log", f"  T2 output: {result['t2']['output_file']}"))
                            self.ui_queue.put(("log", f"  T2 JSON: {result['t2']['output_json']}"))
                        else:
                            self.ui_queue.put(("log", "  T2: skippades"))

                        if result.get("diffusion") is not None:
                            self.ui_queue.put(("log", f"  {DIFFUSION_LABEL}: {diffusion_status}"))
                            self.ui_queue.put(("log", f"  {DIFFUSION_LABEL} output: {result['diffusion']['output_file']}"))
                            self.ui_queue.put(("log", f"  {DIFFUSION_LABEL} JSON: {result['diffusion']['output_json']}"))
                        else:
                            self.ui_queue.put(("log", f"  {DIFFUSION_LABEL}: skippades"))

                        if result.get("perfusion") is not None:
                            self.ui_queue.put(("log", f"  {PERFUSION_LABEL}: {perfusion_status}"))
                            self.ui_queue.put(("log", f"  {PERFUSION_LABEL} output: {result['perfusion']['output_file']}"))
                            self.ui_queue.put(("log", f"  {PERFUSION_LABEL} JSON: {result['perfusion']['output_json']}"))
                        else:
                            self.ui_queue.put(("log", f"  {PERFUSION_LABEL}: skippades"))

                        self.ui_queue.put(("log", ""))

                    except UserCancelledError as exc:
                        session_results.append({
                            "subject": session_dir.parent.name,
                            "session": session_dir.name,
                            "session_dir": str(session_dir),
                            "status": "cancelled",
                            "error": str(exc),
                        })
                        self.ui_queue.put(("log", f"  AVBRUTEN: {exc}"))
                        self.ui_queue.put(("log", ""))

                    except Exception as exc:
                        failed_count += 1
                        error_item = {
                            "subject": session_dir.parent.name,
                            "session": session_dir.name,
                            "session_dir": str(session_dir),
                            "status": "failed",
                            "error": str(exc),
                            "traceback": traceback.format_exc(),
                        }
                        session_results.append(error_item)
                        self.ui_queue.put(("log", f"  FEL: {exc}"))
                        self.ui_queue.put(("log", ""))

                    finally:
                        self.ui_queue.put(("progress", (completed_sessions, total_sessions)))

            self.executor = None

            batch_finished_at = datetime.now()
            batch_elapsed = time.perf_counter() - batch_start_perf
            batch_elapsed_time = format_duration(batch_elapsed)

            similarity_rows = collect_similarity_metric_rows(session_results)
            similarity_metrics_path = summary_root / SIMILARITY_METRICS_FILENAME

            write_similarity_metrics_xlsx(
                path=str(similarity_metrics_path),
                rows=similarity_rows,
            )

            summary = {
                "ProgramName": APP_TITLE,
                "CreatedAt": datetime.now().isoformat(timespec="seconds"),
                "StartedAt": batch_started_at.isoformat(timespec="seconds"),
                "FinishedAt": batch_finished_at.isoformat(timespec="seconds"),
                "ElapsedSeconds": round(batch_elapsed, 3),
                "ElapsedTime": batch_elapsed_time,
                "RootDirectory": str(root_path),
                "BIDSValidation": bids_validation_result,
                "DefaultCostFunction": DEFAULT_COST,
                "DiffusionCostFunction": diffusion_cost,
                "PerfusionCostFunction": perfusion_cost,
                "DefaultDegreesOfFreedom": DEFAULT_DOF,
                "OverwriteExisting": OVERWRITE_EXISTING,
                "UseContrastEnhancedT1T2": use_contrast,
                "RunT2Coregistration": run_t2,
                "RunDiffusionCoregistration": run_diffusion,
                "RunPerfusionCoregistration": run_perfusion,
                "MaxParallelRegistrations": max_workers,
                "TotalSessions": len(all_sessions),
                "ProcessedRegistrations": processed_count,
                "SkippedExistingRegistrations": skipped_count,
                "FailedSessions": failed_count,
                "SimilarityMetricsFile": str(similarity_metrics_path),
                "SimilarityMetricsRows": len(similarity_rows),
                "Results": session_results,
            }

            summary_path = summary_root / "coregistration_batch_summary.json"
            save_json(str(summary_path), summary)

            self.ui_queue.put((
                "done",
                {
                    "summary_path": str(summary_path),
                    "similarity_metrics_path": str(similarity_metrics_path),
                    "processed_count": processed_count,
                    "skipped_count": skipped_count,
                    "failed_count": failed_count,
                    "elapsed_time": batch_elapsed_time,
                    "elapsed_seconds": round(batch_elapsed, 3),
                },
            ))

        except Exception as exc:
            self.ui_queue.put((
                "error",
                {
                    "message": str(exc),
                    "traceback": traceback.format_exc(),
                },
            ))

def start():
    app = CoregBatchApp()
    app.mainloop()


if __name__ == "__main__":
    start()